"""
Baycrest / Apartment Takeoffs format normalizer.
Processes sheet "1 Bldg" with specific column structure.
"""
import re
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook
from app.core.logging import get_logger
from app.services.extraction_stats import ExtractionStats, RowDecision
from app.services.uom_utils import normalize_uom as normalize_uom_canonical

logger = get_logger(__name__)


class BaycrestNormalizer:
    """
    Normalizes Baycrest Excel format takeoff data.

    Column Structure:
    - A: Section headers (e.g., "General", "Corridors")
    - B: Classification/item name (e.g., "Stucco Wall SF", "Unit Doors")
    - C: Main quantity numeric
    - D: Optional secondary numeric (e.g., "Ave SF")
    - E: Notes/text content
    """

    def __init__(self):
        self.current_section = None
        self.header_row = None

    def _norm_sheet_name(self, name: str) -> str:
        """Normalize sheet name: lowercase, strip, collapse whitespace."""
        return " ".join((name or "").strip().lower().split())

    def _find_sheet_by_name(self, workbook, target_name: str) -> Optional[Any]:
        """
        Find a sheet by name using normalized matching.

        Handles variations like 'Units ', ' Units', 'UNITS' all matching 'units'.
        """
        target_norm = self._norm_sheet_name(target_name)

        # Build normalized map
        for sheet_name in workbook.sheetnames:
            if self._norm_sheet_name(sheet_name) == target_norm:
                return workbook[sheet_name]

        # Also try exact match as fallback
        if target_name in workbook.sheetnames:
            return workbook[target_name]

        return None

    def normalize_file(self, file_path: str, target_sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Normalize a Baycrest format Excel file.

        Args:
            file_path: Path to the Excel file
            target_sheet_name: Sheet name to process (if None, uses "1 Bldg" or fallback)

        Returns:
            Dict with raw_rows, raw_data, and extraction stats
        """
        try:
            workbook = load_workbook(file_path, data_only=True)

            # Use provided sheet name or find default
            target_sheet = None
            if target_sheet_name:
                # Try normalized matching first
                target_sheet = self._find_sheet_by_name(workbook, target_sheet_name)
                if target_sheet:
                    logger.info(f"Using provided sheet '{target_sheet.title}'")
                else:
                    logger.warning(f"Sheet '{target_sheet_name}' not found in workbook")

            # Fallback to "1 Bldg" variations
            if not target_sheet:
                target_sheet = self._find_sheet_by_name(workbook, "1 Bldg")
                if target_sheet:
                    logger.info(f"Found target sheet '{target_sheet.title}'")

            # Final fallback to active sheet
            if not target_sheet:
                target_sheet = workbook.active
                logger.warning(f"No matching sheet found, using first sheet: {target_sheet.title}")

            # Process the sheet
            raw_rows = []
            raw_data = []
            stats_tracker = ExtractionStats()

            # Get header row for UOM inference
            if target_sheet.max_row > 0:
                self.header_row = {
                    'C': self._get_cell_value(target_sheet.cell(1, 3)),
                    'D': self._get_cell_value(target_sheet.cell(1, 4))
                }

            # Process all rows
            for row_idx in range(1, target_sheet.max_row + 1):
                row_data = {
                    'row': row_idx,
                    'A': self._get_cell_value(target_sheet.cell(row_idx, 1)),
                    'B': self._get_cell_value(target_sheet.cell(row_idx, 2)),
                    'C': self._get_cell_value(target_sheet.cell(row_idx, 3)),
                    'D': self._get_cell_value(target_sheet.cell(row_idx, 4)),
                    'E': self._get_cell_value(target_sheet.cell(row_idx, 5))
                }

                # Add to raw_rows for audit - ALWAYS add every row we see
                raw_rows.append(row_data)

                # Create decision for this row
                decision = RowDecision(status="EXTRACTED")  # Default to extracted

                # Check if row is completely empty
                is_empty = all(v is None or (isinstance(v, str) and not v.strip())
                              for v in [row_data['A'], row_data['B'], row_data['C'], row_data['D'], row_data['E']])

                if is_empty:
                    decision.status = "IGNORED"
                    decision.add_reason("empty")
                    stats_tracker.commit_row(decision)
                    continue

                # Check for section header in column A
                if row_data['A'] and isinstance(row_data['A'], str) and row_data['A'].strip():
                    potential_section = row_data['A'].strip()
                    # Update current section if this looks like a section header
                    if not any(skip in potential_section.lower() for skip in ['sheet', 'schedule', 'notes']):
                        self.current_section = potential_section
                        logger.debug(f"Updated section to: {self.current_section}")

                    # If no classification in B, this is just a section header row
                    if not row_data['B'] or not isinstance(row_data['B'], str) or not row_data['B'].strip():
                        decision.status = "IGNORED"
                        decision.add_reason("section_header")
                        stats_tracker.commit_row(decision)
                        continue

                # Check if B has classification
                has_classification = (row_data['B'] and
                                    isinstance(row_data['B'], str) and
                                    row_data['B'].strip())

                if not has_classification:
                    decision.status = "IGNORED"
                    decision.add_reason("no_classification")
                    stats_tracker.commit_row(decision)
                    continue

                classification = row_data['B'].strip()

                # Check if C has numeric value
                c_value = self._get_numeric_value(row_data['C'])
                if c_value is None:
                    decision.status = "IGNORED"
                    decision.add_reason("no_quantity")
                    stats_tracker.commit_row(decision)
                    continue

                # Row will be extracted - build measures list
                measures = []

                # Primary measure from column C
                c_uom = self._infer_uom(classification, 'C')
                measures.append({
                    'value': c_value,
                    'uom': c_uom,
                    'source': 'C'
                })

                # Optional secondary measure from column D
                d_value = self._get_numeric_value(row_data['D'])
                if d_value is not None:
                    d_uom = self._infer_uom(classification, 'D')
                    measures.append({
                        'value': d_value,
                        'uom': d_uom,
                        'source': 'D'
                    })

                # Build normalized record
                normalized_record = {
                    'section': self.current_section,
                    'classification': classification,
                    'measures': measures,
                    'provenance': {
                        'sheet': target_sheet.title,
                        'row': row_idx
                    }
                }

                # Add notes if present
                if row_data['E'] and isinstance(row_data['E'], str) and row_data['E'].strip():
                    normalized_record['notes'] = row_data['E'].strip()

                raw_data.append(normalized_record)

                # Mark as extracted
                decision.status = "EXTRACTED"
                stats_tracker.commit_row(decision)

            # Get final stats
            stats = stats_tracker.to_dict()

            logger.info(f"Baycrest extraction complete: {stats}")

            return {
                'raw_rows': raw_rows,
                'raw_data': raw_data,
                'stats': stats
            }

        except Exception as e:
            logger.error(f"Error processing Baycrest file: {str(e)}")
            raise

    def _get_cell_value(self, cell) -> Any:
        """Get cell value, handling None and empty strings."""
        if cell is None:
            return None
        value = cell.value
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def _get_numeric_value(self, value: Any) -> Optional[float]:
        """Convert value to float if numeric, otherwise return None."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Try to parse as number
            clean = value.strip().replace(',', '').replace('$', '')
            try:
                return float(clean)
            except ValueError:
                return None
        return None

    def _infer_uom(self, classification: str, column: str) -> Optional[str]:
        """
        Infer UOM based on classification text and column.

        Rules:
        - If classification contains/ends with " SF" => "SF"
        - If classification contains/ends with " LF" => "LF"
        - If classification contains/ends with " Count" => "EA"
        - For column C: Check header row, default to EA if count-like
        - For column D: Check header row for "SF", otherwise null
        """
        classification_lower = classification.lower()

        # Column D should use header-driven inference first.
        # This avoids class-level hints (e.g., "Total SF") incorrectly
        # turning supplemental "Ave SF" values into mapped measures.
        if column == 'D':
            if self.header_row and self.header_row.get('D'):
                header_d = str(self.header_row['D']).lower()
                if 'ave' in header_d and 'total sf' in classification_lower:
                    return None
                if 'sf' in header_d or 'square' in header_d:
                    return "SF"
            return None

        # Check classification text for UOM hints
        if " sf" in classification_lower or classification_lower.endswith("sf"):
            return "SF"
        if " lf" in classification_lower or classification_lower.endswith("lf"):
            return "LF"
        if " count" in classification_lower or classification_lower.endswith("count"):
            return "EA"

        # Column-specific defaults
        if column == 'C':
            # Check if classification suggests square footage (room/area names)
            sf_keywords = ['lobby', 'lounge', 'fitness', 'guardhouse', 'gaurdhouse', 'clubhouse',
                          'amenities', 'amenity', 'residence services', 'mail room', 'storage',
                          'flooring', 'ceiling', 'wall sf', 'deck', 'vestibule', 'vest sf',
                          'wall subtract', 'subtract', 'rec room', 'garage lid']
            if any(kw in classification_lower for kw in sf_keywords):
                return "SF"

            # Check header row
            if self.header_row and self.header_row.get('C'):
                header_c = str(self.header_row['C']).lower()
                if 'count' in header_c:
                    return "EA"

            # Check if classification suggests counting
            count_keywords = ['door', 'unit', 'fixture', 'outlet', 'switch', 'window', 'vanity', 'toilet']
            if any(kw in classification_lower for kw in count_keywords):
                return "EA"

            # Default for numeric without clear UOM
            return None

        return None

    def normalize_uom(self, uom: str) -> str:
        """
        Normalize UOM strings to canonical format.
        Uses centralized normalize_uom function for consistency.
        Canonical set: EA, SF, LF, LVL (never FT).
        """
        if not uom:
            return uom

        # Use centralized normalization
        normalized = normalize_uom_canonical(uom)
        return normalized if normalized else uom.upper().strip()
