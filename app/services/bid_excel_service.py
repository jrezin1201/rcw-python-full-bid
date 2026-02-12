"""
Excel import/export helpers for bid workflows.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.ui.viewmodels import BidFormState, LineItem, ToggleMask, ProjectInfo

INTERNAL_MARKER = "__RCW_INTERNAL_BID_V1__"

# -- Shared proposal styles --
_CAMBRIA = "Cambria"
_CURRENCY_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_DATE_FMT = "mm-dd-yy"
_THIN = Side(style="thin")

_font_label = Font(name=_CAMBRIA, bold=True, size=11, color="000000")
_font_normal = Font(name=_CAMBRIA, bold=False, size=11, color="000000")
_font_section = Font(name=_CAMBRIA, bold=True, size=12, color="000000")
_font_exclusion = Font(name=_CAMBRIA, italic=True, size=10, color="FF0000")
_font_pricing_hdr = Font(name=_CAMBRIA, bold=True, size=14, color="000000")
_font_bold = Font(name=_CAMBRIA, bold=True, size=11, color="000000")
_font_red_bold = Font(name=_CAMBRIA, bold=True, size=12, color="FF0000")
_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_align_left = Alignment(horizontal="left", vertical="center")
_align_right = Alignment(horizontal="right", vertical="center")
_align_center = Alignment(horizontal="center", vertical="center")


def is_internal_bid_workbook(file_path: str) -> bool:
    """Detect whether a workbook is an editable internal bid export."""
    wb = load_workbook(file_path, data_only=False)
    try:
        if "_rcw_data" in wb.sheetnames:
            ws = wb["_rcw_data"]
            return ws["A1"].value == INTERNAL_MARKER
        return False
    finally:
        wb.close()


def export_internal_bid_workbook(state: BidFormState) -> bytes:
    """
    Export a professional internal workbook with section headers, per-row
    pricing, exclusions, and a pricing summary table.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Bid"

    # -- Styles --
    CURRENCY_FMT = '"$"#,##0.00'
    dark_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    label_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    exclusion_font = Font(italic=True, color="FF0000", size=10)
    bold_font = Font(bold=True, size=10)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # -- Column widths --
    col_widths = {"A": 40, "B": 10, "C": 8, "D": 12, "E": 10, "F": 10, "G": 10, "H": 15}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # -- Header block (rows 1-9) matching proposal layout --
    info = state.project_info
    today = datetime.now(timezone.utc).date().isoformat()

    # Row 1
    ws["A1"] = "Developer:"
    ws["A1"].font = label_font
    ws["B1"] = info.developer or ""
    ws["E1"] = "Date:"
    ws["E1"].font = label_font
    ws["F1"] = today

    # Row 2
    ws["A2"] = "Address:"
    ws["A2"].font = label_font
    ws["B2"] = info.address or ""
    ws["E2"] = "Contact:"
    ws["E2"].font = label_font
    ws["F2"] = info.contact or ""

    # Row 3
    ws["A3"] = "City:"
    ws["A3"].font = label_font
    ws["B3"] = info.city or ""
    ws["E3"] = "Phone:"
    ws["E3"].font = label_font
    ws["F3"] = info.phone or ""

    # Row 4
    ws["E4"] = "Email:"
    ws["E4"].font = label_font
    ws["F4"] = info.email or ""

    # Row 6 - project details
    ws["A6"] = "PROJECT"
    ws["A6"].font = label_font
    ws["B6"] = state.project_name or ""
    ws["E6"] = "PLANS"
    ws["E6"].font = label_font
    ws["G6"] = "DATED"
    ws["G6"].font = label_font

    # Row 7
    unit_count = int(
        sum(
            item.qty
            for item in state.raw_items
            if not item.excluded
            and "unit" in item.name.lower()
            and "count" in item.name.lower()
        )
    )
    total_sf = sum(
        item.qty for item in state.raw_items if not item.excluded and item.uom.upper() == "SF"
    )
    ws["A7"] = "UNITS"
    ws["A7"].font = label_font
    ws["B7"] = f"{unit_count} Units" if unit_count else ""
    ws["E7"] = "ARCH"
    ws["E7"].font = label_font
    ws["G7"] = info.arch_date or ""

    # Row 8
    ws["A8"] = "CITY"
    ws["A8"].font = label_font
    ws["B8"] = info.project_city or info.city or ""
    ws["E8"] = "LANDSCAPE"
    ws["E8"].font = label_font
    ws["G8"] = info.landscape_date or ""

    # Row 9
    ws["A9"] = "SF"
    ws["A9"].font = label_font
    ws["B9"] = round(total_sf, 2) if total_sf else ""
    ws["E9"] = "9900 SPEC"
    ws["E9"].font = label_font
    ws["G9"] = "N/A"

    # -- Helper to write a dark section-header row --
    def write_section_header(row: int, title: str, subtotal: float | None = None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row, 1)
        cell.value = title.upper()
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = left_align
        # Apply fill to all merged columns
        for c in range(2, 8):
            ws.cell(row, c).fill = dark_fill
        # Subtotal in column H
        h_cell = ws.cell(row, 8)
        h_cell.fill = dark_fill
        if subtotal is not None:
            h_cell.value = round(subtotal, 2)
            h_cell.number_format = CURRENCY_FMT
            h_cell.font = header_font
            h_cell.alignment = right_align
        else:
            h_cell.font = header_font

    # -- Scope sections --
    row = 12  # start after header block + blank row

    section_totals_map = {s.section_name.lower(): s.total for s in state.section_totals}

    for section_name in state.get_raw_sections():
        section_items = state.get_raw_items_by_section(section_name)
        active = [i for i in section_items if not i.excluded and not i.is_exclusion and i.qty > 0 and not i.is_alternate]
        exclusions = [i for i in section_items if i.is_exclusion]
        subtotal = section_totals_map.get(section_name.lower(), 0.0)

        # Section header row
        write_section_header(row, section_name, subtotal)
        row += 1

        # Column sub-headers
        for col, label in [(1, "Item"), (2, "Qty"), (3, "UOM"), (4, "$/Unit"), (8, "Total")]:
            c = ws.cell(row, col)
            c.value = label
            c.font = bold_font
            c.alignment = right_align if col >= 2 else left_align
        row += 1

        # Active item rows
        for item in active:
            ws.cell(row, 1).value = item.name
            ws.cell(row, 1).font = normal_font

            ws.cell(row, 2).value = float(item.qty)
            ws.cell(row, 2).font = normal_font

            ws.cell(row, 3).value = item.uom
            ws.cell(row, 3).font = normal_font

            ws.cell(row, 4).value = round(float(item.unit_price_effective), 2)
            ws.cell(row, 4).number_format = CURRENCY_FMT
            ws.cell(row, 4).font = normal_font

            ws.cell(row, 8).value = round(float(item.row_total), 2)
            ws.cell(row, 8).number_format = CURRENCY_FMT
            ws.cell(row, 8).font = normal_font

            row += 1

        # Exclusion rows
        for item in exclusions:
            ws.cell(row, 1).value = f"Excludes {item.name}"
            ws.cell(row, 1).font = exclusion_font
            row += 1

        # Blank spacer
        row += 1

    # -- Pricing Summary Table --
    write_section_header(row, "PRICING")
    row += 1

    # Column headers
    ws.cell(row, 1).value = "Section"
    ws.cell(row, 1).font = bold_font
    ws.cell(row, 8).value = "Amount"
    ws.cell(row, 8).font = bold_font
    ws.cell(row, 8).alignment = right_align
    row += 1

    for section_name in state.get_raw_sections():
        subtotal = section_totals_map.get(section_name.lower(), 0.0)
        ws.cell(row, 1).value = section_name
        ws.cell(row, 1).font = normal_font
        ws.cell(row, 8).value = round(subtotal, 2)
        ws.cell(row, 8).number_format = CURRENCY_FMT
        ws.cell(row, 8).font = normal_font
        row += 1

    # Grand total
    ws.cell(row, 1).value = "Total"
    ws.cell(row, 1).font = bold_font
    ws.cell(row, 8).value = round(float(state.grand_total), 2)
    ws.cell(row, 8).number_format = CURRENCY_FMT
    ws.cell(row, 8).font = bold_font
    ws.cell(row, 8).alignment = right_align
    row += 2

    # -- Alternates section --
    alternates = [i for i in state.raw_items if i.is_alternate]
    if alternates:
        write_section_header(row, "ADD ALTERNATES")
        row += 1

        for item in alternates:
            ws.cell(row, 1).value = item.name
            ws.cell(row, 1).font = normal_font
            ws.cell(row, 8).value = round(float(item.row_total), 2)
            ws.cell(row, 8).number_format = CURRENCY_FMT
            ws.cell(row, 8).font = normal_font
            row += 1

    # -- Hidden data sheet for round-trip import --
    _write_data_sheet(wb, state)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_data_sheet(wb: Workbook, state: BidFormState) -> None:
    """Write a hidden '_rcw_data' sheet with full structured data for re-import."""
    ds = wb.create_sheet("_rcw_data")
    ds.sheet_state = "hidden"

    # Row 1: marker
    ds["A1"] = INTERNAL_MARKER

    # Row 2: project name
    ds["A2"] = "project_name"
    ds["B2"] = state.project_name or ""

    # Rows 3-11: project info
    info = state.project_info
    for r, (key, val) in enumerate([
        ("developer", info.developer),
        ("address", info.address),
        ("city", info.city),
        ("contact", info.contact),
        ("phone", info.phone),
        ("email", info.email),
        ("project_city", info.project_city),
        ("arch_date", info.arch_date),
        ("landscape_date", info.landscape_date),
    ], start=3):
        ds.cell(r, 1).value = key
        ds.cell(r, 2).value = val or ""

    # Row 13: column headers
    headers = [
        "section", "name", "qty", "uom", "base_price", "difficulty",
        "add_1", "add_2", "add_3", "add_4", "add_5",
        "tax", "labor", "materials", "equipment", "subcontractor",
        "mult", "excluded", "is_exclusion", "notes", "is_alternate",
    ]
    for c, h in enumerate(headers, start=1):
        ds.cell(13, c).value = h

    # Row 14+: item data
    for r, item in enumerate(state.raw_items, start=14):
        ds.cell(r, 1).value = item.section
        ds.cell(r, 2).value = item.name
        ds.cell(r, 3).value = float(item.qty)
        ds.cell(r, 4).value = item.uom
        ds.cell(r, 5).value = float(item.unit_price_base)
        ds.cell(r, 6).value = item.difficulty
        for lvl in range(1, 6):
            ds.cell(r, 6 + lvl).value = float(item.difficulty_adders.get(lvl, 0.0))
        ds.cell(r, 12).value = item.toggle_mask.tax
        ds.cell(r, 13).value = item.toggle_mask.labor
        ds.cell(r, 14).value = item.toggle_mask.materials
        ds.cell(r, 15).value = item.toggle_mask.equipment
        ds.cell(r, 16).value = item.toggle_mask.subcontractor
        ds.cell(r, 17).value = float(item.mult)
        ds.cell(r, 18).value = item.excluded
        ds.cell(r, 19).value = item.is_exclusion
        ds.cell(r, 20).value = item.notes or ""
        ds.cell(r, 21).value = item.is_alternate


def import_internal_bid_workbook(file_path: str) -> BidFormState:
    """Import state from an editable internal workbook export."""
    wb = load_workbook(file_path, data_only=False)
    try:
        if "_rcw_data" not in wb.sheetnames:
            raise ValueError("Not a supported internal workbook format")
        ds = wb["_rcw_data"]
        if ds["A1"].value != INTERNAL_MARKER:
            raise ValueError("Not a supported internal workbook format")

        project_name = _string(ds["B2"].value) or Path(file_path).stem

        # Read project info from rows 3-11
        info_fields = {}
        for r in range(3, 12):
            key = _string(ds.cell(r, 1).value)
            val = _string(ds.cell(r, 2).value)
            if key:
                info_fields[key] = val

        project_info = ProjectInfo(
            developer=info_fields.get("developer"),
            address=info_fields.get("address"),
            city=info_fields.get("city"),
            contact=info_fields.get("contact"),
            phone=info_fields.get("phone"),
            email=info_fields.get("email"),
            project_city=info_fields.get("project_city"),
            arch_date=info_fields.get("arch_date"),
            landscape_date=info_fields.get("landscape_date"),
        )

        # Read items from row 14+
        items: list[LineItem] = []
        row = 14
        while True:
            section = ds.cell(row, 1).value
            name = ds.cell(row, 2).value
            if not section and not name:
                break

            qty = _float(ds.cell(row, 3).value)
            uom = _string(ds.cell(row, 4).value) or "EA"
            base_price = _float(ds.cell(row, 5).value)
            difficulty = int(_float(ds.cell(row, 6).value) or 1)
            adders = {
                1: _float(ds.cell(row, 7).value),
                2: _float(ds.cell(row, 8).value),
                3: _float(ds.cell(row, 9).value),
                4: _float(ds.cell(row, 10).value),
                5: _float(ds.cell(row, 11).value),
            }

            toggle_mask = ToggleMask(
                tax=_bool(ds.cell(row, 12).value, True),
                labor=_bool(ds.cell(row, 13).value, True),
                materials=_bool(ds.cell(row, 14).value, True),
                equipment=_bool(ds.cell(row, 15).value, False),
                subcontractor=_bool(ds.cell(row, 16).value, False),
            )
            mult = _float(ds.cell(row, 17).value, 1.0)
            excluded = _bool(ds.cell(row, 18).value, False)
            is_exclusion = _bool(ds.cell(row, 19).value, False)
            notes = _string(ds.cell(row, 20).value)
            is_alt = _bool(ds.cell(row, 21).value, False)

            items.append(
                LineItem(
                    id=f"imported_{row}",
                    section=_string(section) or "General",
                    name=_string(name) or f"Item {row}",
                    qty=qty,
                    uom=uom,
                    unit_price_base=base_price,
                    difficulty=max(1, min(5, difficulty)),
                    difficulty_adders=adders,
                    toggle_mask=toggle_mask,
                    mult=mult,
                    excluded=excluded,
                    is_exclusion=is_exclusion,
                    notes=notes,
                    is_alternate=is_alt,
                )
            )
            row += 1

        return BidFormState(
            project_name=project_name,
            raw_items=items,
            created_at=datetime.now(timezone.utc).isoformat(),
            source_file=Path(file_path).name,
            project_info=project_info,
        )
    finally:
        wb.close()


def export_proposal_workbook(state: BidFormState) -> bytes:
    """
    Export client-facing proposal workbook built from scratch (no template).
    Layout mirrors the reference proposal XLSX.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # -- Column widths --
    for letter, w in [("A", 10.5), ("B", 10), ("C", 15), ("D", 16),
                      ("E", 15.5), ("F", 14), ("G", 9), ("H", 9)]:
        ws.column_dimensions[letter].width = w

    info = state.project_info
    today = datetime.now(timezone.utc).date()

    # Derived metrics
    unit_count = int(sum(
        i.qty for i in state.raw_items
        if not i.excluded and "unit" in i.name.lower() and "count" in i.name.lower()
    ))
    total_sf = sum(
        i.qty for i in state.raw_items if not i.excluded and i.uom.upper() == "SF"
    )

    # ── Helper: write merged row ──
    def _merged_text(row: int, text: str, font=_font_normal, align=_align_left):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row, 1)
        c.value = text
        c.font = font
        c.alignment = align

    # ── HEADER (rows 1-4) ──
    def _hdr(row, col, val, font=_font_label):
        c = ws.cell(row, col)
        c.value = val
        c.font = font

    _hdr(1, 1, "Developer:")
    ws.merge_cells("B1:D1")
    ws.cell(1, 2).value = info.developer or ""
    ws.cell(1, 2).font = _font_normal
    ws.cell(1, 2).alignment = _align_left
    _hdr(1, 5, "Date:")
    ws.merge_cells("F1:H1")
    ws.cell(1, 6).value = today
    ws.cell(1, 6).font = _font_normal
    ws.cell(1, 6).number_format = _DATE_FMT
    ws.cell(1, 6).alignment = _align_right

    _hdr(2, 1, "Address:")
    ws.merge_cells("B2:D2")
    ws.cell(2, 2).value = info.address or ""
    ws.cell(2, 2).font = _font_normal
    ws.cell(2, 2).alignment = _align_left
    _hdr(2, 5, "Contact: ")
    ws.merge_cells("F2:H2")
    ws.cell(2, 6).value = info.contact or ""
    ws.cell(2, 6).font = _font_bold
    ws.cell(2, 6).alignment = _align_right

    _hdr(3, 1, "City:")
    ws.merge_cells("B3:D3")
    ws.cell(3, 2).value = info.city or ""
    ws.cell(3, 2).font = _font_normal
    ws.cell(3, 2).alignment = _align_left
    _hdr(3, 5, "Phone:")
    ws.merge_cells("F3:H3")
    ws.cell(3, 6).value = info.phone or ""
    ws.cell(3, 6).font = _font_bold
    ws.cell(3, 6).alignment = _align_right

    _hdr(4, 5, "Email: ")
    ws.merge_cells("F4:H4")
    ws.cell(4, 6).value = info.email or ""
    ws.cell(4, 6).font = _font_normal
    ws.cell(4, 6).alignment = _align_right

    # ── PROJECT INFO (rows 6-10) ──
    _hdr(6, 1, "PROJECT")
    ws.merge_cells("B6:D6")
    ws.cell(6, 2).value = state.project_name or ""
    ws.cell(6, 2).font = _font_normal
    ws.cell(6, 2).alignment = _align_left
    ws.merge_cells("E6:F6")
    ws.cell(6, 5).value = "PLANS"
    ws.cell(6, 5).font = _font_label
    ws.cell(6, 5).alignment = _align_center
    ws.merge_cells("G6:H6")
    ws.cell(6, 7).value = "DATED"
    ws.cell(6, 7).font = _font_label
    ws.cell(6, 7).alignment = _align_center

    _hdr(7, 1, "UNITS")
    ws.merge_cells("B7:D7")
    ws.cell(7, 2).value = f"{unit_count} Residential Units" if unit_count else ""
    ws.cell(7, 2).font = _font_normal
    ws.cell(7, 2).alignment = _align_left
    ws.merge_cells("E7:F7")
    ws.cell(7, 5).value = "ARCHITECTURAL"
    ws.cell(7, 5).font = _font_normal
    ws.cell(7, 5).alignment = _align_left
    ws.merge_cells("G7:H7")
    ws.cell(7, 7).value = info.arch_date or ""
    ws.cell(7, 7).font = _font_normal
    ws.cell(7, 7).alignment = _align_center

    _hdr(8, 1, "CITY")
    ws.merge_cells("B8:D8")
    ws.cell(8, 2).value = info.project_city or ""
    ws.cell(8, 2).font = _font_normal
    ws.cell(8, 2).alignment = _align_left
    ws.merge_cells("E8:F8")
    ws.cell(8, 5).value = "LANDSCAPE"
    ws.cell(8, 5).font = _font_normal
    ws.cell(8, 5).alignment = _align_left
    ws.merge_cells("G8:H8")
    landscape_val = info.landscape_date or "Excluded"
    ws.cell(8, 7).value = landscape_val
    ws.cell(8, 7).font = Font(name=_CAMBRIA, size=11, color="FF0000") if landscape_val == "Excluded" else _font_normal
    ws.cell(8, 7).alignment = _align_center

    ws.merge_cells("E9:F9")
    ws.cell(9, 5).value = "INTERIOR DESIGN"
    ws.cell(9, 5).font = _font_normal
    ws.cell(9, 5).alignment = _align_left
    ws.merge_cells("G9:H9")
    ws.cell(9, 7).value = "NA"
    ws.cell(9, 7).font = _font_normal
    ws.cell(9, 7).alignment = _align_center

    ws.merge_cells("E10:F10")
    ws.cell(10, 5).value = "OWNER SPECS"
    ws.cell(10, 5).font = Font(name=_CAMBRIA, size=10)
    ws.cell(10, 5).alignment = _align_left
    ws.merge_cells("G10:H10")
    ws.cell(10, 7).value = "NA"
    ws.cell(10, 7).font = _font_normal
    ws.cell(10, 7).alignment = _align_center

    # ── SCOPE SECTIONS (dynamic rows starting at 12) ──
    row = 12
    section_totals_map = {s.section_name.lower(): s.total for s in state.section_totals}

    for section_name in state.get_raw_sections():
        section_items = state.get_raw_items_by_section(section_name)
        active = [i for i in section_items if not i.excluded and not i.is_exclusion and i.qty > 0 and not i.is_alternate]
        exclusions = [i for i in section_items if i.is_exclusion]

        if not active and not exclusions:
            continue

        # Section header
        ws.cell(row, 1).value = section_name.upper()
        ws.cell(row, 1).font = _font_section
        ws.row_dimensions[row].height = 15.6
        row += 1

        # Active scope items (merged A:H)
        for item in active:
            _merged_text(row, item.name)
            row += 1

        # Exclusions
        for item in exclusions:
            ws.cell(row, 1).value = f"Excludes {item.name}"
            ws.cell(row, 1).font = _font_exclusion
            ws.cell(row, 1).alignment = _align_left
            row += 1

        # Blank spacer
        row += 1

    # ── PRICING SECTION ──
    pricing_start = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).value = "PRICING"
    ws.cell(row, 1).font = _font_pricing_hdr
    ws.cell(row, 1).alignment = _align_center
    ws.row_dimensions[row].height = 17.4
    # Borders on pricing header
    for c in range(1, 9):
        cell = ws.cell(row, c)
        cell.border = Border(
            left=_THIN if c == 1 else None,
            right=_THIN if c == 8 else None,
        )
    row += 1

    # Amount header
    ws.cell(row, 5).value = "Amount"
    ws.cell(row, 5).font = _font_bold
    ws.cell(row, 5).alignment = _align_center
    row += 1

    # Section pricing rows
    for section_name in state.get_raw_sections():
        subtotal = section_totals_map.get(section_name.lower(), 0.0)
        if subtotal <= 0:
            continue
        ws.cell(row, 4).value = section_name.capitalize()
        ws.cell(row, 4).font = _font_normal
        ws.cell(row, 5).value = round(subtotal, 2)
        ws.cell(row, 5).number_format = _CURRENCY_FMT
        ws.cell(row, 5).font = _font_normal
        row += 1

    # Total row
    ws.cell(row, 4).value = "Total"
    ws.cell(row, 4).font = _font_bold
    ws.cell(row, 5).value = round(float(state.grand_total), 2)
    ws.cell(row, 5).number_format = _CURRENCY_FMT
    ws.cell(row, 5).font = _font_bold
    row += 1

    # Net wrap note
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row, 1).value = "Pricing Net Wrap Liability Insurance"
    ws.cell(row, 1).font = Font(name=_CAMBRIA, italic=True, size=11, color="FF0000")
    ws.cell(row, 1).alignment = _align_center
    row += 2

    # ── EXCLUSIONS LIST ──
    ws.cell(row, 4).value = "EXCLUSIONS"
    ws.cell(row, 4).font = _font_red_bold
    row += 1

    exclusion_pairs = [
        ("9900 Specifications", "Masked Hinges"),
        ("0 VOC Paints & Systems", "ALL Stand Pipes"),
        ("Scaffolding & Lifts", "Wall Coverings"),
        ("Saturday & Weekend Work", "Paid Parking: Parking To Be Provided By Contractor"),
        ("Water Proofing Membranes", "Caulking Windows: By Drywall Contractor"),
        ("ALL Exterior Caulking: Done By Other", "Not Responsible For Rusting If Metal Is Not Metalized"),
        ("Power & Pressure Washing", "Payment & Performance Bonds"),
        ("Signing Unmodified Scaffold Agreements", "Excess & Umbrella Coverages"),
    ]
    for left, right in exclusion_pairs:
        ws.cell(row, 1).value = left
        ws.cell(row, 1).font = _font_normal
        ws.cell(row, 5).value = right
        ws.cell(row, 5).font = _font_normal
        row += 1
    # Highlight 9900 spec
    spec_row = row - len(exclusion_pairs)
    ws.cell(spec_row, 1).font = _font_bold
    ws.cell(spec_row, 1).fill = _fill_yellow

    row += 1

    # ── ADD ALTERNATES ──
    alternates = [i for i in state.raw_items if i.is_alternate]
    if alternates:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).value = "ADD ALTERNATES"
        ws.cell(row, 1).font = _font_bold
        ws.cell(row, 1).alignment = _align_center
        # Bottom border
        for c in range(1, 7):
            ws.cell(row, c).border = Border(bottom=_THIN)
        row += 1

        for item in alternates:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row, 1).value = item.name
            ws.cell(row, 1).font = _font_normal
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            ws.cell(row, 4).value = round(float(item.row_total), 2)
            ws.cell(row, 4).number_format = _CURRENCY_FMT
            ws.cell(row, 4).font = _font_normal
            row += 1

        row += 1

    # ── ADDITIONAL WORK RATES ──
    ws.cell(row, 1).value = "ADDITIONAL WORK CHARGED AT:"
    ws.cell(row, 1).font = _font_bold
    ws.cell(row, 4).value = "$73.00/HR"
    ws.cell(row, 4).font = _font_normal
    ws.cell(row, 4).alignment = _align_right
    row += 1

    ws.cell(row, 1).value = "1/2 Time OT Work"
    ws.cell(row, 1).font = _font_bold
    ws.cell(row, 4).value = "$37.00/HR"
    ws.cell(row, 4).font = _font_normal
    ws.cell(row, 4).alignment = _align_right
    row += 2

    # ── MATERIALS ──
    ws.cell(row, 1).value = "MATERIALS INCLUDED IN BID: VISTA PAINTS"
    ws.cell(row, 1).font = _font_section
    row += 1

    mat_sections = [
        ("Units", [("Flat", "Breezewall"), ("Enamel", "V-Pro")]),
        ("Common Area", [("Flat", "Breezewall"), ("Enamel", "V-Pro")]),
        ("Exterior", [("Doors", "Protec"), ("Balcony Rails", "Protec"), ("Wood", "Coverall")]),
    ]
    for mat_title, items in mat_sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1).value = mat_title
        ws.cell(row, 1).font = _font_section
        ws.cell(row, 1).alignment = _align_left
        row += 1
        for label, product in items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row, 1).value = label
            ws.cell(row, 1).font = _font_normal
            ws.cell(row, 1).alignment = _align_left
            ws.cell(row, 3).value = product
            ws.cell(row, 3).font = _font_normal
            row += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _float(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _string(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _bool(value, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return default


