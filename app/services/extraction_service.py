"""
Excel extraction service for processing spreadsheet files.
"""
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging import get_logger
from app.schemas.job import ExtractionResult, QAReport

logger = get_logger(__name__)


class ExcelExtractor:
    """
    Excel file extraction service with robust handling for:
    - Merged cells
    - Header detection
    - Type inference
    - Data quality validation
    """

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.extracted_rows: List[Dict[str, Any]] = []
        self.qa_warnings: List[Dict[str, Any]] = []
        self.columns: List[str] = []
        self.unmapped_columns: List[str] = []
        self.type_anomalies: List[Dict[str, Any]] = []
        self.suspected_totals: List[int] = []
        self.empty_rows_count = 0

    def extract(self, sheet_name: Optional[str] = None) -> Tuple[ExtractionResult, QAReport]:
        """
        Main extraction method.
        Returns extracted data and QA report.
        """
        try:
            # Load workbook
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,  # Get calculated values, not formulas
                read_only=False  # Need write access for merged cell handling
            )

            # Select worksheet
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")
                worksheet = self.workbook[sheet_name]
            else:
                worksheet = self.workbook.active

            # Extract data from worksheet
            self._extract_worksheet(worksheet)

            # Create result
            result = ExtractionResult(
                rows=self.extracted_rows,
                columns=self.columns,
                provenance={
                    "file": str(self.file_path.name),
                    "sheet": worksheet.title,
                    "extracted_at": datetime.now(timezone.utc).isoformat(),
                    "total_sheets": len(self.workbook.sheetnames)
                }
            )

            # Create QA report
            qa = self._generate_qa_report()

            return result, qa

        except Exception as e:
            logger.error(f"Extraction failed: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()

    def _extract_worksheet(self, worksheet: Worksheet):
        """Extract data from a single worksheet."""
        if worksheet.max_row == 0:
            logger.warning(f"Empty worksheet: {worksheet.title}")
            return

        # Get merged cell ranges for proper value expansion
        merged_ranges = list(worksheet.merged_cells.ranges)

        # Find header row
        header_row_idx = self._detect_header_row(worksheet)
        if not header_row_idx:
            logger.warning("No header row detected, using first row")
            header_row_idx = 1

        # Extract headers
        headers = self._extract_headers(worksheet, header_row_idx)
        self.columns = headers

        # Extract data rows
        for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            row_data = self._extract_row(
                worksheet,
                row_idx,
                headers,
                merged_ranges
            )

            if row_data:
                # Check if this might be a totals row
                if self._is_suspected_total_row(row_data, row_idx):
                    self.suspected_totals.append(row_idx)

                # Add provenance
                row_data["__provenance"] = {
                    "sheet_name": worksheet.title,
                    "excel_row_index": row_idx,
                    "source_cell_range": f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}"
                }

                self.extracted_rows.append(row_data)
            else:
                self.empty_rows_count += 1

    def _detect_header_row(self, worksheet: Worksheet) -> Optional[int]:
        """
        Detect the header row by looking for:
        - Row with most non-empty cells
        - Row with text values (not numbers)
        - Common header keywords
        """
        header_keywords = {
            'name', 'description', 'quantity', 'qty', 'price',
            'amount', 'total', 'unit', 'cost', 'item', 'product',
            'code', 'part', 'number', 'date', 'id'
        }

        best_row = None
        best_score = 0

        for row_idx in range(1, min(20, worksheet.max_row + 1)):  # Check first 20 rows
            score = 0
            non_empty = 0

            for cell in worksheet[row_idx]:
                if cell.value:
                    non_empty += 1
                    value_str = str(cell.value).lower().strip()

                    # Check for header keywords
                    if any(keyword in value_str for keyword in header_keywords):
                        score += 2

                    # Prefer text over numbers for headers
                    if isinstance(cell.value, str):
                        score += 1

            # Weight by number of non-empty cells
            if non_empty > 0:
                score = score * non_empty

            if score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    def _extract_headers(self, worksheet: Worksheet, row_idx: int) -> List[str]:
        """Extract and normalize header names."""
        headers = []
        seen_headers = {}

        for col_idx, cell in enumerate(worksheet[row_idx], 1):
            if cell.value:
                # Normalize header name
                header = self._normalize_header(str(cell.value))

                # Handle duplicate headers
                if header in seen_headers:
                    seen_headers[header] += 1
                    header = f"{header}_{seen_headers[header]}"
                else:
                    seen_headers[header] = 1

                headers.append(header)
            else:
                # Handle empty header
                headers.append(f"column_{col_idx}")
                self.unmapped_columns.append(f"column_{col_idx}")

        return headers

    def _normalize_header(self, header: str) -> str:
        """Normalize header names for consistency."""
        # Convert to lowercase
        header = header.lower().strip()

        # Remove special characters, keep alphanumeric and underscores
        header = re.sub(r'[^\w\s]', '', header)

        # Replace spaces with underscores
        header = re.sub(r'\s+', '_', header)

        # Remove leading/trailing underscores
        header = header.strip('_')

        return header or "unnamed"

    def _extract_row(
        self,
        worksheet: Worksheet,
        row_idx: int,
        headers: List[str],
        merged_ranges: List
    ) -> Optional[Dict[str, Any]]:
        """Extract data from a single row."""
        row_data = {}
        has_data = False

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = self._get_cell_value(cell, merged_ranges)

            if value is not None:
                has_data = True

                # Store both raw and display value when different
                row_data[header] = value

                # Check for type anomalies
                self._check_type_anomaly(header, value, row_idx, col_idx)

        return row_data if has_data else None

    def _get_cell_value(self, cell: Cell, merged_ranges: List) -> Any:
        """
        Get the value of a cell, handling merged cells properly.
        For merged cells, returns the value from the top-left cell of the range.
        """
        # Check if cell is part of a merged range
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                # Get the top-left cell of the merged range
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                master_cell = cell.parent.cell(row=min_row, column=min_col)
                return master_cell.value

        # Regular cell
        return cell.value

    def _check_type_anomaly(self, header: str, value: Any, row_idx: int, col_idx: int):
        """Check for type anomalies in data."""
        header_lower = header.lower()

        # Check for quantity/amount fields that should be numeric
        if any(keyword in header_lower for keyword in ['qty', 'quantity', 'amount', 'count']):
            if isinstance(value, str) and value.strip():
                # Check if it's not a valid number
                try:
                    float(value.replace(',', ''))
                except ValueError:
                    self.type_anomalies.append({
                        "type": "non_numeric_quantity",
                        "header": header,
                        "value": value,
                        "row": row_idx,
                        "column": col_idx,
                        "message": f"Expected numeric value for '{header}', got text: '{value}'"
                    })

        # Check for price/cost fields
        if any(keyword in header_lower for keyword in ['price', 'cost', 'rate']):
            if isinstance(value, str) and '%' in value:
                self.type_anomalies.append({
                    "type": "percentage_in_price",
                    "header": header,
                    "value": value,
                    "row": row_idx,
                    "column": col_idx,
                    "message": f"Price field '{header}' contains percentage: '{value}'"
                })

        # Check for scientific notation that might lose precision
        if isinstance(value, (int, float)) and abs(value) > 1e10:
            self.type_anomalies.append({
                "type": "scientific_notation_risk",
                "header": header,
                "value": value,
                "row": row_idx,
                "column": col_idx,
                "message": f"Large number may be displayed in scientific notation: {value}"
            })

    def _is_suspected_total_row(self, row_data: Dict[str, Any], row_idx: int) -> bool:
        """
        Detect if a row might be a totals/summary row.
        """
        total_indicators = ['total', 'subtotal', 'sum', 'grand total', 'summary']

        for key, value in row_data.items():
            if value and isinstance(value, str):
                value_lower = value.lower().strip()
                if any(indicator in value_lower for indicator in total_indicators):
                    return True

        return False

    def _generate_qa_report(self) -> QAReport:
        """Generate quality assurance report."""
        # Calculate confidence score
        confidence = self._calculate_confidence()

        # Compile all warnings
        warnings = []

        if self.unmapped_columns:
            warnings.append({
                "type": "unmapped_columns",
                "severity": "warning",
                "columns": self.unmapped_columns,
                "message": f"Found {len(self.unmapped_columns)} columns without clear headers"
            })

        if self.suspected_totals:
            warnings.append({
                "type": "suspected_totals",
                "severity": "info",
                "rows": self.suspected_totals,
                "message": f"Found {len(self.suspected_totals)} potential total/summary rows"
            })

        for anomaly in self.type_anomalies:
            warnings.append({
                "type": "type_anomaly",
                "severity": "warning",
                **anomaly
            })

        return QAReport(
            rows_extracted=len(self.extracted_rows),
            unmapped_columns=self.unmapped_columns,
            empty_rows_removed=self.empty_rows_count,
            suspected_totals_rows=self.suspected_totals,
            type_anomalies=self.type_anomalies,
            confidence=confidence,
            warnings=warnings
        )

    def _calculate_confidence(self) -> float:
        """
        Calculate confidence score based on extraction quality.
        Score between 0.0 and 1.0.
        """
        if not self.extracted_rows:
            return 0.0

        score = 1.0

        # Deduct for unmapped columns
        if self.unmapped_columns:
            unmapped_ratio = len(self.unmapped_columns) / len(self.columns) if self.columns else 1
            score -= unmapped_ratio * 0.2

        # Deduct for type anomalies
        if self.type_anomalies:
            anomaly_ratio = min(len(self.type_anomalies) / len(self.extracted_rows), 1.0)
            score -= anomaly_ratio * 0.3

        # Deduct for too many empty rows
        if self.empty_rows_count > len(self.extracted_rows):
            score -= 0.1

        # Deduct for suspected totals mixed with data
        if self.suspected_totals:
            total_ratio = len(self.suspected_totals) / len(self.extracted_rows)
            score -= total_ratio * 0.1

        return max(score, 0.0)


# TODO: PDF Extraction Service
class PDFExtractor:
    """
    Placeholder for PDF extraction service.

    Phase 2 Implementation:
    - Use pypdf2 or pdfplumber for text extraction
    - Use tabula-py for table extraction
    - OCR support with pytesseract for scanned PDFs
    - Layout analysis for complex documents
    """
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def extract(self) -> Tuple[ExtractionResult, QAReport]:
        raise NotImplementedError("PDF extraction will be implemented in Phase 2")
