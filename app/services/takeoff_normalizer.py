"""
Excel normalization service for takeoff files.
Handles the specific format: Classification | Quantity | Quantity1 UOM | Quantity2 | Quantity2 UOM | Quantity3 | Quantity3 UOM
"""
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging import get_logger

logger = get_logger(__name__)


class TakeoffNormalizer:
    """
    Normalizes takeoff Excel files into a standard format.
    """

    # Expected column patterns with more synonyms (case-insensitive)
    COLUMN_PATTERNS = {
        'classification': re.compile(
            r'classification|class|description|item|name|product|material|work.?item|scope',
            re.IGNORECASE
        ),
        'quantity': re.compile(
            r'^quantity$|^qty$|^quantity1$|^qty1$|^quantity.?1$|^takeoff.?qty$|^amount$|^count$',
            re.IGNORECASE
        ),
        'quantity_uom': re.compile(
            r'quantity.*uom|qty.*uom|quantity1.*uom|uom1|^uom$|^unit$|^units$|quantity.?1.?uom|qty.?1.?uom',
            re.IGNORECASE
        ),
        'quantity2': re.compile(
            r'quantity.?2|qty.?2|quantity2|qty2|second.?qty',
            re.IGNORECASE
        ),
        'quantity2_uom': re.compile(
            r'quantity.?2.*uom|qty.?2.*uom|uom.?2|quantity2.*uom|qty2.*uom',
            re.IGNORECASE
        ),
        'quantity3': re.compile(
            r'quantity.?3|qty.?3|quantity3|qty3|third.?qty',
            re.IGNORECASE
        ),
        'quantity3_uom': re.compile(
            r'quantity.?3.*uom|qty.?3.*uom|uom.?3|quantity3.*uom|qty3.*uom',
            re.IGNORECASE
        ),
    }

    # UOM normalization mappings
    UOM_MAPPINGS = {
        'FT': 'LF',
        'FEET': 'LF',
        'LINEAR FEET': 'LF',
        'LINEAR FT': 'LF',
        'LIN FT': 'LF',
        'SQFT': 'SF',
        'SQ FT': 'SF',
        'SQ.FT.': 'SF',
        'SQUARE FEET': 'SF',
        'EACH': 'EA',
        'PCS': 'EA',
        'PIECES': 'EA',
        'PIECE': 'EA',
        'UNIT': 'EA',
        'UNITS': 'EA',
        'COUNT': 'EA'
    }

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.column_mapping = {}
        self.header_row_index = None
        self.rows_ignored = 0  # Track ignored rows (headers, blanks, totals)

    def parse_excel_to_normalized_rows(self, sheet_name: Optional[str] = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """
        Parse Excel file and return normalized rows.

        Returns:
            Tuple of (normalized_rows, metadata)
            metadata includes: rows_ignored count
        """
        try:
            # Load workbook
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,  # Get calculated values, not formulas
                read_only=False
            )

            logger.info(f"Loaded workbook with sheets: {self.workbook.sheetnames}")

            # Select worksheet - try to find one with headers if not specified
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                worksheet = self.workbook[sheet_name]
                logger.info(f"Using specified sheet: {sheet_name}")
            else:
                # Try each sheet to find one with valid headers
                worksheet = None
                for sheet in self.workbook.worksheets:
                    self._detect_columns(sheet)
                    if 'classification' in self.column_mapping:
                        worksheet = sheet
                        logger.info(f"Auto-selected sheet with headers: {sheet.title}")
                        break

                if worksheet is None:
                    # Fall back to active sheet if no headers found
                    worksheet = self.workbook.active
                    logger.warning(f"No sheet with headers found, using active: {worksheet.title}")
                    self._detect_columns(worksheet)

            # Create comprehensive debug summary
            debug_summary = self._create_debug_summary(worksheet)
            logger.info("=" * 80)
            logger.info("EXTRACTION DEBUG SUMMARY")
            logger.info("=" * 80)
            for line in debug_summary.split('\n'):
                if line.strip():
                    logger.info(line)
            logger.info("=" * 80)

            # Extract and normalize rows
            normalized_rows = self._extract_rows(worksheet)

            # Log first few extracted rows for debugging
            if normalized_rows:
                logger.info("First 5 extracted rows:")
                for i, row in enumerate(normalized_rows[:5]):
                    measures_str = ", ".join([f"{m['value']} {m['uom']}" for m in row['measures']])
                    logger.info(f"  Row {i+1}: '{row['classification']}' -> {measures_str or 'no measures'}")
            else:
                logger.warning("NO ROWS EXTRACTED - Check column mapping and data format")

            # Return rows and metadata
            metadata = {
                'rows_ignored': self.rows_ignored,
                'sheet_name': worksheet.title,
                'total_rows_processed': worksheet.max_row,
                'header_row_index': self.header_row_index,
                'column_mapping': self._format_column_mapping(),
                'rows_extracted': len(normalized_rows),
                'debug_summary': debug_summary
            }

            logger.info(f"Extraction complete: {len(normalized_rows)} rows extracted, {self.rows_ignored} ignored")

            return normalized_rows, metadata

        except Exception as e:
            logger.error(f"Failed to parse Excel: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()

    def _detect_columns(self, worksheet: Worksheet):
        """
        Detect header row and map columns to expected fields.
        """
        # Reset for this sheet
        self.column_mapping = {}
        self.header_row_index = None

        # Check first 50 rows for headers
        max_scan_rows = min(50, worksheet.max_row + 1)
        best_mapping = {}
        best_row_idx = None
        best_score = 0

        for row_idx in range(1, max_scan_rows):
            row_values = []
            for cell in worksheet[row_idx]:
                if cell.value:
                    # Normalize: lowercase, strip, remove punctuation
                    value = str(cell.value).strip()
                    row_values.append(value)
                else:
                    row_values.append('')

            # Try to match columns
            mapping = self._match_columns(row_values)

            # Score the mapping (more matches = better)
            score = len(mapping)

            # Bonus points if we have classification and at least one quantity
            if 'classification' in mapping:
                score += 2
            if 'quantity' in mapping:
                score += 1
            if 'quantity_uom' in mapping:
                score += 1

            if score > best_score:
                best_score = score
                best_mapping = mapping
                best_row_idx = row_idx

                # If we found a really good match, stop early
                if 'classification' in mapping and 'quantity' in mapping:
                    break

        # Use the best mapping found
        if best_row_idx and best_mapping:
            self.header_row_index = best_row_idx
            self.column_mapping = best_mapping

            # Log the actual header values for debugging
            logger.info(f"Found header row at index {best_row_idx} with score {best_score}")
            self._log_header_details(worksheet, best_row_idx)
        else:
            # If no header found, assume first row and try to map
            logger.warning("No clear header row found, using row 1")
            self.header_row_index = 1
            first_row = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
            self.column_mapping = self._match_columns(first_row)
            self._log_header_details(worksheet, 1)

    def _match_columns(self, headers: List[str]) -> Dict[str, int]:
        """
        Match column headers to expected fields.
        Returns mapping of field_name -> column_index (0-based)
        """
        mapping = {}

        # Log raw headers for debugging
        logger.debug(f"Matching headers: {headers[:12]}")

        # First, try standard takeoff format with position-based detection
        # Expected: Classification | Quantity1 | UOM1 | Quantity2 | UOM2 | Quantity3 | UOM3
        if len(headers) >= 7:
            # Check if this looks like the standard alternating pattern
            # Column 0 should be classification-like
            # Odd columns (1,3,5) should be quantities
            # Even columns (2,4,6) should be UOMs

            is_standard_format = True

            # Check if first column looks like classification
            if headers[0]:
                h0_lower = headers[0].strip().lower()
                if not any(word in h0_lower for word in ['class', 'description', 'item', 'name', 'material', 'scope']):
                    is_standard_format = False

            # Check if odd columns look like quantities
            for idx in [1, 3, 5]:
                if idx < len(headers) and headers[idx]:
                    h_lower = headers[idx].strip().lower()
                    if not any(word in h_lower for word in ['quantity', 'qty', 'amount', 'count']) and not h_lower.replace('.','').replace(',','').replace(' ','').isdigit():
                        is_standard_format = False
                        break

            # Check if even columns after 0 look like UOMs
            for idx in [2, 4, 6]:
                if idx < len(headers) and headers[idx]:
                    h_lower = headers[idx].strip().lower()
                    # UOM columns often have unit names or "uom"
                    if not any(word in h_lower for word in ['uom', 'unit', 'ea', 'sf', 'lf', 'each', 'sqft']):
                        is_standard_format = False
                        break

            if is_standard_format:
                logger.info("Detected standard takeoff format with alternating columns")
                # Map using position
                mapping['classification'] = 0
                if len(headers) > 1: mapping['quantity'] = 1
                if len(headers) > 2: mapping['quantity_uom'] = 2
                if len(headers) > 3: mapping['quantity2'] = 3
                if len(headers) > 4: mapping['quantity2_uom'] = 4
                if len(headers) > 5: mapping['quantity3'] = 5
                if len(headers) > 6: mapping['quantity3_uom'] = 6

                logger.debug(f"Position-based mapping: {mapping}")
                return mapping

        # Fallback: Pattern-based detection for non-standard formats
        logger.info("Using pattern-based column detection")

        for idx, header in enumerate(headers):
            if not header:
                continue

            # Normalize header: lowercase, strip spaces
            header_clean = str(header).strip().lower()

            # Try to match each expected column
            matched = False

            # Classification/Description (highest priority)
            if not matched and idx == 0:  # First column is usually classification
                if any(word in header_clean for word in ['class', 'description', 'item', 'name', 'material', 'scope', 'work']):
                    mapping['classification'] = idx
                    matched = True

            # Check for quantity columns with specific numbers
            if not matched:
                if '3' in header_clean and ('quantity' in header_clean or 'qty' in header_clean):
                    if 'uom' in header_clean or 'unit' in header_clean:
                        if 'quantity3_uom' not in mapping:
                            mapping['quantity3_uom'] = idx
                            matched = True
                    elif 'quantity3' not in mapping:
                        mapping['quantity3'] = idx
                        matched = True
                elif '2' in header_clean and ('quantity' in header_clean or 'qty' in header_clean):
                    if 'uom' in header_clean or 'unit' in header_clean:
                        if 'quantity2_uom' not in mapping:
                            mapping['quantity2_uom'] = idx
                            matched = True
                    elif 'quantity2' not in mapping:
                        mapping['quantity2'] = idx
                        matched = True
                elif ('1' in header_clean or header_clean in ['quantity', 'qty']) and ('quantity' in header_clean or 'qty' in header_clean):
                    if 'uom' in header_clean or 'unit' in header_clean:
                        if 'quantity_uom' not in mapping:
                            mapping['quantity_uom'] = idx
                            matched = True
                    elif 'quantity' not in mapping:
                        mapping['quantity'] = idx
                        matched = True

            # Check for standalone UOM columns
            if not matched and ('uom' in header_clean or 'unit' in header_clean or header_clean in ['ea', 'sf', 'lf', 'each', 'sqft']):
                # Try to determine which UOM based on position relative to quantities
                if idx > 0:
                    # Check if previous column is a quantity
                    if 'quantity3' in mapping and mapping['quantity3'] == idx - 1:
                        mapping['quantity3_uom'] = idx
                        matched = True
                    elif 'quantity2' in mapping and mapping['quantity2'] == idx - 1:
                        mapping['quantity2_uom'] = idx
                        matched = True
                    elif 'quantity' in mapping and mapping['quantity'] == idx - 1:
                        mapping['quantity_uom'] = idx
                        matched = True

                # If still not matched, use number hints or default assignment
                if not matched:
                    if '3' in header_clean and 'quantity3_uom' not in mapping:
                        mapping['quantity3_uom'] = idx
                        matched = True
                    elif '2' in header_clean and 'quantity2_uom' not in mapping:
                        mapping['quantity2_uom'] = idx
                        matched = True
                    elif 'quantity_uom' not in mapping:
                        mapping['quantity_uom'] = idx
                        matched = True

            # Final fallback for classification
            if not matched and 'classification' not in mapping:
                for field_name, pattern in self.COLUMN_PATTERNS.items():
                    if field_name == 'classification' and pattern.search(header_clean):
                        mapping['classification'] = idx
                        matched = True
                        break

        # If no classification found, assume first column
        if 'classification' not in mapping and len(headers) > 0 and headers[0]:
            logger.warning("No classification column found, using first column")
            mapping['classification'] = 0

        logger.debug(f"Pattern-based mapping result: {mapping}")
        return mapping

    def _log_header_details(self, worksheet, row_idx):
        """Log the actual header cell values for debugging."""
        headers = []
        for col_idx, cell in enumerate(worksheet[row_idx][:12], 0):  # First 12 columns
            value = str(cell.value) if cell.value else '(empty)'
            headers.append(f"Col{col_idx}={value}")

        logger.info(f"Header row {row_idx} cells: {' | '.join(headers)}")

    def _format_column_mapping(self):
        """Format column mapping for logging."""
        if not self.column_mapping:
            return "{}"

        formatted = {}
        for field, idx in self.column_mapping.items():
            formatted[f"{field}(col{idx})"] = idx
        return formatted

    def _create_debug_summary(self, worksheet) -> str:
        """Create a comprehensive debug summary for the extraction."""
        lines = []

        lines.append(f"Sheet: {worksheet.title}")
        lines.append(f"Total rows in sheet: {worksheet.max_row}")
        lines.append(f"Header row index: {self.header_row_index}")
        lines.append("")

        # Show actual header values
        lines.append("Header cells (first 12 columns):")
        if self.header_row_index:
            for col_idx, cell in enumerate(worksheet[self.header_row_index][:12]):
                value = str(cell.value) if cell.value else '(empty)'
                lines.append(f"  Col {col_idx}: {value}")
        lines.append("")

        # Show column mapping
        lines.append("Column mapping detected:")
        if self.column_mapping:
            for field, idx in sorted(self.column_mapping.items(), key=lambda x: x[1]):
                if idx < 12 and self.header_row_index:
                    header_value = worksheet[self.header_row_index][idx].value
                    lines.append(f"  {field:20s} -> Col {idx}: '{header_value}'")
                else:
                    lines.append(f"  {field:20s} -> Col {idx}")
        else:
            lines.append("  NO MAPPING DETECTED!")
        lines.append("")

        # Show first 5 data rows raw values
        lines.append("First 5 data rows (raw values):")
        start_row = (self.header_row_index or 0) + 1
        end_row = min(start_row + 5, worksheet.max_row + 1)

        for row_idx in range(start_row, end_row):
            row_values = []
            for col_idx in range(min(7, worksheet.max_column)):
                cell = worksheet[row_idx][col_idx]
                value = str(cell.value) if cell.value else '(empty)'
                if len(value) > 20:
                    value = value[:20] + '...'
                row_values.append(value)
            lines.append(f"  Row {row_idx}: {' | '.join(row_values)}")

        return '\n'.join(lines)

    def _extract_rows(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Extract and normalize data rows.
        """
        normalized_rows = []

        # Start from row after header
        start_row = (self.header_row_index or 0) + 1
        self.rows_ignored = self.header_row_index or 0  # Count header rows as ignored

        for row_idx in range(start_row, worksheet.max_row + 1):
            row_cells = list(worksheet[row_idx])

            # Extract classification
            classification = self._get_cell_value(row_cells, 'classification')
            if not classification:
                self.rows_ignored += 1  # Blank row
                continue  # Skip rows without classification

            # Detect total/subtotal rows
            if self._is_total_row(classification):
                self.rows_ignored += 1  # Total/subtotal row
                logger.debug(f"Ignoring total row: {classification}")
                continue

            # Extract measures
            measures = []

            # Quantity 1
            qty1 = self._get_numeric_value(row_cells, 'quantity')
            uom1 = self._get_cell_value(row_cells, 'quantity_uom')
            if qty1 is not None and uom1:
                measures.append({
                    'value': qty1,
                    'uom': self._normalize_uom(uom1),
                    'source': 'Quantity'
                })

            # Quantity 2
            qty2 = self._get_numeric_value(row_cells, 'quantity2')
            uom2 = self._get_cell_value(row_cells, 'quantity2_uom')
            if qty2 is not None and uom2:
                measures.append({
                    'value': qty2,
                    'uom': self._normalize_uom(uom2),
                    'source': 'Quantity2'
                })

            # Quantity 3
            qty3 = self._get_numeric_value(row_cells, 'quantity3')
            uom3 = self._get_cell_value(row_cells, 'quantity3_uom')
            if qty3 is not None and uom3:
                measures.append({
                    'value': qty3,
                    'uom': self._normalize_uom(uom3),
                    'source': 'Quantity3'
                })

            # Only add row if it has measures
            if measures:
                normalized_rows.append({
                    'classification': classification.strip(),
                    'measures': measures,
                    'provenance': {
                        'sheet': worksheet.title,
                        'row': row_idx
                    }
                })

        logger.info(f"Extracted {len(normalized_rows)} normalized rows")
        return normalized_rows

    def _get_cell_value(self, row_cells: List, field_name: str) -> Optional[str]:
        """Get string value from cell based on field mapping."""
        if field_name not in self.column_mapping:
            return None

        idx = self.column_mapping[field_name]
        if idx >= len(row_cells):
            return None

        value = row_cells[idx].value
        if value is None:
            return None

        return str(value).strip()

    def _get_numeric_value(self, row_cells: List, field_name: str) -> Optional[float]:
        """
        Get numeric value from cell, handling commas and conversions.
        ONLY use this for quantity columns, never for UOM columns.
        """
        # Safety check: Never parse UOM columns as numeric
        if 'uom' in field_name.lower():
            logger.error(f"Attempted to parse UOM column '{field_name}' as numeric - this is a bug!")
            return None

        value_str = self._get_cell_value(row_cells, field_name)
        if not value_str:
            return None

        try:
            # Remove commas and spaces
            cleaned = value_str.replace(',', '').replace(' ', '')

            # Try to convert to float
            return float(cleaned)
        except (ValueError, AttributeError):
            # Not a valid number
            logger.debug(f"Could not convert '{value_str}' to number for field '{field_name}'")
            return None

    def _normalize_uom(self, uom: str) -> str:
        """
        Normalize unit of measure to standard format.
        """
        if not uom:
            return ''

        uom_upper = uom.upper().strip()

        # Apply mappings
        if uom_upper in self.UOM_MAPPINGS:
            return self.UOM_MAPPINGS[uom_upper]

        # Return as-is if no mapping found
        return uom_upper

    def _is_total_row(self, classification: str) -> bool:
        """
        Detect if a classification indicates a total/subtotal row.
        """
        total_indicators = [
            'total', 'subtotal', 'sub-total', 'grand total',
            'sum', 'summary', 'aggregate'
        ]

        class_lower = classification.lower().strip()

        # Check if the entire classification is just a total indicator
        if class_lower in total_indicators:
            return True

        # Check if it starts with a total indicator
        for indicator in total_indicators:
            if class_lower.startswith(indicator + ' ') or class_lower.startswith(indicator + ':'):
                return True

        return False