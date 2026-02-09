from __future__ import annotations

from dataclasses import dataclass, field
from typing import List, Optional, Tuple, Dict, Any
from openpyxl import load_workbook


@dataclass
class SheetSelection:
    """Details about how the takeoff sheet was selected."""
    selected_sheet: Optional[str]
    method: str  # "exact" | "prefix" | "score" | "none"
    candidates_tried: List[str] = field(default_factory=list)
    score: Optional[float] = None


@dataclass
class SignatureCheck:
    ok: bool
    score: float
    matched_sheet: Optional[str]
    warnings: List[str]
    debug: Dict[str, Any]
    sheet_selection: Optional[SheetSelection] = None


# Known Baycrest section headers (column A)
BAYCREST_SECTION_HEADERS = {
    "general",
    "corridors",
    "exterior",
    "units",
    "stairs",
    "amenity",
    "garage",
    "landscape",
}

# Default takeoff sheet prefix
DEFAULT_TAKEOFF_PREFIX = "1 bldg"


def norm_sheet_name(name: str) -> str:
    """Normalize sheet name: lowercase, strip, collapse whitespace."""
    return " ".join((name or "").strip().lower().split())


def _norm(v: Any) -> str:
    """Normalize cell value to lowercase string."""
    if v is None:
        return ""
    return str(v).strip().lower()


def _is_numeric_like(value: Any) -> bool:
    """Check if a value is numeric or can be parsed as numeric."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        clean = value.strip().replace(',', '').replace('$', '')
        try:
            float(clean)
            return True
        except ValueError:
            return False
    return False


def _is_label_string(value: Any) -> bool:
    """Check if value is a non-empty label string (not purely numeric)."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return False
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return False
        # Must not be purely numeric
        try:
            float(s.replace(',', '').replace('$', ''))
            return False
        except ValueError:
            return True
    return False


def _score_sheet_content(ws, max_rows: int = 200) -> Tuple[int, int, float]:
    """
    Score a sheet by Baycrest content signals.

    Returns:
        (section_hits, data_rows, score)
    """
    section_hits = 0
    data_rows = 0

    for row_idx in range(1, min(ws.max_row + 1, max_rows + 1)):
        # Check column A for section headers
        col_a = _norm(ws.cell(row_idx, 1).value)
        if col_a in BAYCREST_SECTION_HEADERS:
            section_hits += 1

        # Check for data-like rows: B has label, C is numeric
        col_b = ws.cell(row_idx, 2).value
        col_c = ws.cell(row_idx, 3).value

        if _is_label_string(col_b) and _is_numeric_like(col_c):
            data_rows += 1

    # Score formula: section headers weighted more heavily
    score = (section_hits * 2) + min(data_rows, 100) / 10

    return section_hits, data_rows, score


def _select_takeoff_sheet(wb, sheet_map: Dict[str, str]) -> SheetSelection:
    """
    Select the takeoff sheet using prioritized rules.

    Rules:
    1. Exact match for normalized "1 bldg"
    2. Prefix match for sheets starting with "1 bldg"
    3. Fallback to score-by-content
    """
    candidates_tried = []

    # Rule 1: Exact match for "1 bldg"
    if DEFAULT_TAKEOFF_PREFIX in sheet_map:
        original_name = sheet_map[DEFAULT_TAKEOFF_PREFIX]
        return SheetSelection(
            selected_sheet=original_name,
            method="exact",
            candidates_tried=[original_name]
        )

    # Rule 2: Prefix match for sheets starting with "1 bldg"
    prefix_matches = []
    for norm_name, original_name in sheet_map.items():
        if norm_name.startswith(DEFAULT_TAKEOFF_PREFIX):
            prefix_matches.append(original_name)
            candidates_tried.append(original_name)

    if prefix_matches:
        # Pick first match (alphabetically sorted for consistency)
        prefix_matches.sort()
        return SheetSelection(
            selected_sheet=prefix_matches[0],
            method="prefix",
            candidates_tried=prefix_matches
        )

    # Rule 3: Fallback to score-by-content
    best_sheet = None
    best_score = 0.0
    best_section_hits = 0
    best_data_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_hits, data_rows, score = _score_sheet_content(ws)
        candidates_tried.append(f"{sheet_name} (score={score:.1f})")

        # Minimum thresholds to accept
        if section_hits >= 3 and data_rows >= 15:
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                best_section_hits = section_hits
                best_data_rows = data_rows

    if best_sheet:
        return SheetSelection(
            selected_sheet=best_sheet,
            method="score",
            candidates_tried=candidates_tried,
            score=best_score
        )

    # No suitable sheet found
    return SheetSelection(
        selected_sheet=None,
        method="none",
        candidates_tried=candidates_tried
    )


def validate_baycrest_workbook(xlsx_path: str) -> SignatureCheck:
    """
    Validate a workbook as Baycrest format.

    Detection is content-based:
    - Select takeoff sheet by name pattern or content scoring
    - Units and Bid Form sheets are OPTIONAL (warnings only)
    - Baycrest is detected by section headers in column A and data rows
    """
    wb = load_workbook(xlsx_path, data_only=True)

    warnings: List[str] = []
    debug: Dict[str, Any] = {"sheets": wb.sheetnames}

    # Build normalized sheet map for matching
    sheet_map = {norm_sheet_name(n): n for n in wb.sheetnames}
    debug["sheet_map"] = sheet_map

    # 1) Check for optional sheets (warnings only, not failures)
    if "units" not in sheet_map:
        warnings.append("Missing 'Units' sheet (optional; Units may exist as a section inside the takeoff sheet).")
    if "bid form" not in sheet_map:
        warnings.append("Missing 'Bid Form' sheet (optional).")

    # 2) Select the takeoff sheet
    sheet_selection = _select_takeoff_sheet(wb, sheet_map)
    debug["sheet_selection"] = {
        "selected_sheet": sheet_selection.selected_sheet,
        "method": sheet_selection.method,
        "candidates_tried": sheet_selection.candidates_tried,
        "score": sheet_selection.score
    }

    # 3) Validate content if sheet was found
    ok = False
    score = 0.0
    matched_sheet = sheet_selection.selected_sheet

    if matched_sheet:
        ws = wb[matched_sheet]
        section_hits, data_rows, content_score = _score_sheet_content(ws)

        debug["content_validation"] = {
            "section_hits": section_hits,
            "data_rows": data_rows,
            "content_score": content_score
        }

        # Pass if we have enough Baycrest signals
        # Either: selected by name (exact/prefix) OR selected by score (already passed thresholds)
        if sheet_selection.method in ("exact", "prefix"):
            # Name-based selection: validate content loosely
            # At least 1 section header and 5 data rows
            ok = section_hits >= 1 and data_rows >= 5
            if not ok:
                warnings.append(
                    f"Sheet '{matched_sheet}' selected by name but has insufficient Baycrest content "
                    f"(section_hits={section_hits}, data_rows={data_rows})."
                )
        elif sheet_selection.method == "score":
            # Score-based selection already passed thresholds
            ok = True

        score = content_score
    else:
        warnings.append(
            f"Could not find a suitable takeoff sheet. "
            f"Tried: {sheet_selection.candidates_tried}"
        )

    return SignatureCheck(
        ok=ok,
        score=score,
        matched_sheet=matched_sheet,
        warnings=warnings,
        debug=debug,
        sheet_selection=sheet_selection,
    )