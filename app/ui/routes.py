"""
UI routes for the Bid Form application.
Handles page rendering and HTMX partial updates.
"""

import os
import re
import uuid
import tempfile
from pathlib import Path
from typing import Optional
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, File, Form, Request, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates

from app.core.logging import get_logger
from app.ui.viewmodels import BidFormState, LineItem, ToggleMask, ProjectInfo, SpecItem, MaterialItem
from app.ui.state import get_current_state, set_state, has_current_bid, get_current_warnings, set_warnings, set_debug, get_current_debug
from app.ui.catalog_service import BidCatalog
from app.ui.excel_mapper import map_excel_with_catalog
from app.ui.constants import DIFFICULTY_LEVELS, SECTION_ORDER
from app.services.baycrest_normalizer import BaycrestNormalizer
from app.services.validators.baycrest_signature import validate_baycrest_workbook
from app.services.bid_excel_service import (
    export_internal_bid_workbook,
    export_proposal_workbook,
    import_internal_bid_workbook,
    is_internal_bid_workbook,
)

logger = get_logger(__name__)

# Create router
router = APIRouter()

# Setup templates
templates = Jinja2Templates(directory="app/templates")


def _fmt_date(value) -> str:
    """Format date values for display — strips time, handles datetime objects and strings."""
    if value is None:
        return ""
    from datetime import datetime as _dt, date as _d
    if isinstance(value, _dt):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, _d):
        return value.strftime("%m/%d/%Y")
    s = str(value).strip()
    if not s:
        return ""
    # Strip trailing 00:00:00 from strings like "2023-12-22 00:00:00"
    s = re.sub(r'\s+00:00:00$', '', s)
    # Try to parse ISO date strings into MM/DD/YYYY
    try:
        parsed = _dt.strptime(s, "%Y-%m-%d")
        return parsed.strftime("%m/%d/%Y")
    except ValueError:
        pass
    return s


templates.env.filters["fmt_date"] = _fmt_date

# ========== Helper Functions ==========

def format_currency(value: float) -> str:
    """Format a number as currency."""
    return f"${value:,.2f}"

def format_number(value: float, max_decimals: int = 2) -> str:
    """Format number with comma separators and trimmed decimals."""
    formatted = f"{value:,.{max_decimals}f}"
    if "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted

def format_currency_input(value: float) -> str:
    """Format number for currency input display."""
    return f"{value:,.2f}"

def parse_numeric_input(raw: str, field_name: str) -> float:
    """Parse numeric user input allowing commas and currency symbols."""
    cleaned = str(raw or "").strip().replace(",", "").replace("$", "")
    if cleaned == "":
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}")
    try:
        return float(cleaned)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}")


def get_template_context(request: Request, **kwargs):
    """Get base template context with common data."""
    context = {
        "request": request,
        "format_currency": format_currency,
        "format_number": format_number,
        "format_currency_input": format_currency_input,
        "current_year": datetime.now().year,
        **kwargs
    }

    # Add current bid if available
    if has_current_bid():
        context["has_bid"] = True
        state = get_current_state()
        if state:
            context["project_name"] = state.project_name
            context["total_items"] = state.total_items
            context["source_file"] = state.source_file

    return context


def _project_header_vars(state: BidFormState) -> dict:
    """Compute derived values used by the project info header partial."""
    unit_count = int(sum(
        i.qty for i in state.raw_items
        if not i.excluded and "unit" in i.name.lower() and "count" in i.name.lower()
    ))
    total_sf = sum(
        i.qty for i in state.raw_items if not i.excluded and i.uom.upper() == "SF"
    )
    return {
        "now_date": datetime.now(timezone.utc).date().isoformat(),
        "unit_count": unit_count,
        "total_sf": total_sf,
    }


# ========== Page Routes ==========

@router.get("/", response_class=HTMLResponse)
async def home_page(request: Request):
    """Render the home page with upload form."""
    context = get_template_context(request, page="home")
    return templates.TemplateResponse("index.html", context)


@router.post("/upload")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    template: str = Form(default="baycrest_v1")
):
    """
    Handle file upload, parse Excel, and redirect to bid form.
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")

        # Save uploaded file temporarily
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)

        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)

        logger.info(f"Processing uploaded file: {file.filename}")

        # If this is an editable internal workbook, import it directly.
        if is_internal_bid_workbook(file_path):
            bid_state = import_internal_bid_workbook(file_path)
            qa_warnings = []
            debug_payload = {}
        else:
            # Validate template signature
            if template == "baycrest_v1":
                sig = validate_baycrest_workbook(file_path)
                if not sig.ok:
                    # Clean up
                    os.remove(file_path)
                    os.rmdir(temp_dir)
                    raise HTTPException(
                        status_code=400,
                        detail=f"File doesn't match Baycrest template: {'; '.join(sig.warnings)}"
                    )

            # Parse the Excel file using CATALOG-DRIVEN mapping
            # This ensures UOM comes from catalog, not extraction defaults
            bid_state, qa_warnings, debug_payload = map_excel_with_catalog(file_path, template)

        # Store state
        bid_id = str(uuid.uuid4())
        bid_state.project_id = bid_id
        bid_state.source_file = file.filename
        bid_state.created_at = datetime.now(timezone.utc).isoformat()

        # Store warnings for display in UI
        set_warnings(bid_id, qa_warnings)

        set_state(bid_id, bid_state)
        set_debug(bid_id, debug_payload)

        # Clean up temp file
        os.remove(file_path)
        os.rmdir(temp_dir)

        # Redirect to bid form
        return RedirectResponse(url="/bid", status_code=303)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing upload: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")



def sort_sections(sections: list) -> list:
    """Sort sections by the defined order."""
    def section_sort_key(section_name):
        name_lower = section_name.lower()
        for i, ordered in enumerate(SECTION_ORDER):
            if ordered.lower() == name_lower or ordered.lower() in name_lower or name_lower in ordered.lower():
                return (i, section_name)
        return (len(SECTION_ORDER), section_name)  # Unknown sections go at the end

    return sorted(sections, key=section_sort_key)


@router.get("/bid", response_class=HTMLResponse)
async def bid_form_page(request: Request):
    """Render the bid form page (raw Excel line items)."""
    state = get_current_state()

    if not state:
        from app.ui.excel_mapper import create_sample_bid_form
        sample_state = create_sample_bid_form()
        set_state("sample", sample_state)
        state = sample_state

    raw_sections = state.get_raw_sections()

    context = get_template_context(
        request,
        page="bid",
        bid_state=state,
        raw_sections=raw_sections,
        difficulty_options=DIFFICULTY_LEVELS,
        **_project_header_vars(state),
    )

    return templates.TemplateResponse("bid_form2.html", context)


@router.post("/bid/section/add", response_class=HTMLResponse)
async def add_section(request: Request, name: str = Form(...)):
    """Add a new empty section and return its HTML block."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    section = name.strip()
    if not section:
        raise HTTPException(status_code=400, detail="Section name is required")

    context = get_template_context(
        request,
        section=section,
        priced_count=0,
        section_total=0,
    )
    html = templates.get_template("partials/new_section_block.html").render(context)
    return HTMLResponse(html)


@router.get("/bid/section-header", response_class=HTMLResponse)
async def get_section_header(request: Request, section: str):
    """Return a single section header row for HTMX refresh."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    items = state.get_raw_items_by_section(section)
    active = [i for i in items if not i.excluded]
    priced = [i for i in active if i.qty > 0]

    context = get_template_context(
        request,
        section=section,
        priced_count=len(priced),
        section_total=sum(i.row_total for i in active),
    )
    html = templates.get_template("partials/section_header.html").render(context)
    return HTMLResponse(html)


# ========== HTMX Partial Update Routes ==========

@router.get("/bid/totals", response_class=HTMLResponse)
async def get_totals(request: Request):
    """Return the totals panel HTML for HTMX refresh."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    context = get_template_context(
        request,
        bid_state=state
    )

    totals_html = templates.get_template("partials/totals.html").render(context)
    return HTMLResponse(totals_html)


@router.post("/bid/item/{item_id}/qty", response_class=HTMLResponse)
async def update_item_qty(
    request: Request,
    item_id: str,
    qty: str = Form(...)
):
    """Update item quantity and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    # Update quantity
    parsed_qty = parse_numeric_input(qty, "quantity")
    if not state.update_item_qty(item_id, parsed_qty):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)

    # Return only the row HTML; use HX-Trigger header to refresh totals panel
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/difficulty", response_class=HTMLResponse)
async def update_item_difficulty(
    request: Request,
    item_id: str,
    difficulty: int = Form(...)
):
    """Update item difficulty and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    # Update difficulty
    if not state.set_item_difficulty(item_id, difficulty):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)

    # Return only the row HTML; use HX-Trigger header to refresh totals panel
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/difficulty-add", response_class=HTMLResponse)
async def update_item_difficulty_add(
    request: Request,
    item_id: str,
    level: int = Form(...),
    amount: str = Form(...)
):
    """Update absolute difficulty add-on for an item and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    parsed_amount = parse_numeric_input(amount, "difficulty amount")
    if not state.update_item_difficulty_add(item_id, level, parsed_amount):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)
    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/toggle", response_class=HTMLResponse)
async def toggle_item_flag(
    request: Request,
    item_id: str,
    toggle_name: str = Form(...)
):
    """Toggle an item flag and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    # Toggle the flag
    if not state.toggle_item(item_id, toggle_name):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)

    # Return only the row HTML; use HX-Trigger header to refresh totals panel
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/price", response_class=HTMLResponse)
async def update_item_price(
    request: Request,
    item_id: str,
    unit_price: str = Form(...)
):
    """Update item unit price and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    # Find and update the item
    item = state.get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    parsed_price = parse_numeric_input(unit_price, "unit price")
    item.unit_price_base = max(0, parsed_price)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)

    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/mult", response_class=HTMLResponse)
async def update_item_mult(
    request: Request,
    item_id: str,
    mult: str = Form(...)
):
    """Update item multiplier and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    parsed_mult = parse_numeric_input(mult, "multiplier")
    if not state.update_item_mult(item_id, max(0, parsed_mult)):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/exclude", response_class=HTMLResponse)
async def toggle_item_excluded(
    request: Request,
    item_id: str,
):
    """Toggle excluded (soft-delete) state for an item and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    if not state.toggle_excluded(item_id):
        raise HTTPException(status_code=404, detail="Item not found")

    item = state.get_item(item_id)

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/exclusion", response_class=HTMLResponse)
async def toggle_item_exclusion(request: Request, item_id: str):
    """Toggle whether an item shows as an exclusion on the print proposal."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    item = state.get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    item.is_exclusion = not item.is_exclusion

    context = get_template_context(
        request, item=item, bid_state=state, difficulty_options=DIFFICULTY_LEVELS
    )
    row_html = templates.get_template("partials/bid_row.html").render(context)
    return HTMLResponse(row_html)


@router.post("/bid/item/add", response_class=HTMLResponse)
async def add_item(
    request: Request,
    name: str = Form(...),
    section: str = Form(...),
    qty: str = Form(default="0"),
    uom: str = Form(default="EA"),
    unit_price: str = Form(default="0"),
    target: str = Form(default="items"),
):
    """Add a new custom line item and return its row HTML."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    parsed_qty = parse_numeric_input(qty, "quantity")
    parsed_price = parse_numeric_input(unit_price, "unit price")

    new_item = LineItem(
        section=section,
        name=name.strip(),
        qty=max(0, parsed_qty),
        uom=uom.strip().upper(),
        unit_price_base=max(0, parsed_price),
    )

    if target == "raw_items":
        state.raw_items.append(new_item)
    else:
        state.add_item(new_item)

    context = get_template_context(
        request,
        item=new_item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/notes", response_class=HTMLResponse)
async def update_item_notes(
    request: Request,
    item_id: str,
    notes: str = Form(default="")
):
    """Update item notes and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    item = state.get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Update notes (empty string becomes None)
    item.notes = notes.strip() if notes.strip() else None

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    return HTMLResponse(row_html)


@router.post("/bid/item/{item_id}/name", response_class=HTMLResponse)
async def update_item_name(
    request: Request,
    item_id: str,
    name: str = Form(...)
):
    """Update item name and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    item = state.get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    trimmed = name.strip()
    if trimmed:
        item.name = trimmed

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


@router.post("/bid/item/{item_id}/uom", response_class=HTMLResponse)
async def update_item_uom(
    request: Request,
    item_id: str,
    uom: str = Form(...)
):
    """Update item unit of measure and return updated row."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    item = state.get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    trimmed = uom.strip()
    if trimmed:
        item.uom = trimmed

    context = get_template_context(
        request,
        item=item,
        bid_state=state,
        difficulty_options=DIFFICULTY_LEVELS
    )

    row_html = templates.get_template("partials/bid_row.html").render(context)
    response = HTMLResponse(row_html)
    response.headers["HX-Trigger"] = "totals-updated"
    return response


# ========== Additional Routes ==========

DEFAULT_SPEC_ITEMS = [
    "Doors, Base & Casings: Semi Gloss Same Color As Walls",
    "Wardrobe & WIC: Walls, Ceilings, Shelving: ALL Same Color As Walls Semi Gloss Finish",
    "Bathroom Walls and Ceilings: Semi Gloss Same Color As Living",
    "Masking Hinges",
    "Two Tone In Living Area",
    "Cased Windows",
    "Base Over Flooring (See Add Alts)",
    "Accent Walls At Living Area",
    "Stool & Apron",
    "Prime Coat At Walls (To be painted 2 coats of flat)",
    "Prep Coat",
]

DEFAULT_EXCLUSIONS = [
    "9900 Specifications",
    "Masked Hinges",
    "0 VOC Paints & Systems",
    "ALL Stand Pipes",
    "Scaffolding & Lifts",
    "Wall Coverings",
    "Saturday & Weekend Work",
    "Paid Parking: Parking To Be Provided By Contractor",
    "Water Proofing Membranes",
    "Caulking Windows: By Drywall Contractor",
    "ALL Exterior Caulking: Done By Other",
    "Not Responsible For Rusting If Metal Is Not Metalized",
    "Power & Pressure Washing",
    "Payment & Performance Bonds",
    "Signing Unmodified Scaffold Agreements",
    "Excess & Umbrella Coverages",
]


# Pricing for spec items when shown as Add Alts (keyed by "item name|section name")
SPEC_ALT_PRICING = {
    "True Prime Coat|Units": 3500.00,
    "Eggshell Walls|Units": 8400.00,
    "Two Tone|Units": 5950.00,
    "Base Over Floor|Units": 1260.00,
    "Mask Hinges|Units": 3000.00,
    "Smooth Wall|Units": 6300.00,
    "True Prime Coat|Corridor & Stairwells": 3240.00,
    "Eggshell Walls|Corridor & Stairwells": 5400.00,
    "Two Tone|Corridor & Stairwells": 3645.00,
    "Smooth Wall|Corridor & Stairwells": 3375.00,
    "Stucco Body|Exterior": 9690.00,
    "Prime Coat at Stucco Body|Exterior": 5100.00,
    "Stucco Accents at Balc|Exterior": 2750.00,
    "Prime Stucco Accents at Balc|Exterior": 2000.00,
    "Balcony Swing Doors|Exterior": 1250.00,
    "Garage Walls Flat|Garage": 2850.00,
    "Garage Columns|Garage": 320.00,
}


def _match_alt_price(item_name: str, section_name: str) -> Optional[float]:
    """Find a price for a spec item by matching against SPEC_ALT_PRICING keys.

    Matches bidirectionally: pricing key name in item name, OR item name in
    pricing key name. Section match is preferred but not required.
    """
    # Exact key match first
    key = f"{item_name}|{section_name}"
    if key in SPEC_ALT_PRICING:
        return SPEC_ALT_PRICING[key]
    # Bidirectional name matching with section preference
    name_lower = item_name.lower()
    section_lower = section_name.lower()
    best_match = None
    best_has_section = False
    for pricing_key, price in SPEC_ALT_PRICING.items():
        pk_name, pk_section = pricing_key.split("|", 1)
        pk_name_lower = pk_name.lower()
        # Check name match in either direction
        name_hit = pk_name_lower in name_lower or name_lower in pk_name_lower
        if not name_hit:
            # Try matching significant words (3+ chars)
            pk_words = {w for w in pk_name_lower.split() if len(w) >= 3}
            name_words = {w for w in name_lower.split() if len(w) >= 3}
            common = pk_words & name_words
            name_hit = len(common) >= 2 or (len(common) >= 1 and len(pk_words) == 1)
        if not name_hit:
            continue
        section_hit = pk_section.lower() in section_lower or section_lower in pk_section.lower()
        # Prefer section+name match over name-only
        if section_hit and not best_has_section:
            best_match = price
            best_has_section = True
        elif section_hit == best_has_section and best_match is None:
            best_match = price
    return best_match


DEFAULT_MATERIALS = {
    "Units": [("Flat", False), ("Enamel", False)],
    "Common Area": [("Flat", False), ("Enamel", False)],
    "Exterior": [
        ("Stucco Body:", False), ("Canopies", False),
        ("Balcony Rails", False), ("Doors", False),
        ("Wood", False), ("Concrete Walls", False),
    ],
}
DEFAULT_MATERIALS_SECTION_ORDER = ["Units", "Common Area", "Exterior"]


def _ensure_spec_items(state: BidFormState) -> None:
    """Initialize default spec items and exclusions for any section that doesn't have them yet."""
    for st in state.section_totals:
        if st.section_name not in state.spec_items:
            state.spec_items[st.section_name] = [
                SpecItem(name=n) for n in DEFAULT_SPEC_ITEMS
            ]
        # Populate prices for items that don't have one yet
        for spec in state.spec_items[st.section_name]:
            if spec.price is None:
                spec.price = _match_alt_price(spec.name, st.section_name)
    if not state.spec_exclusions:
        state.spec_exclusions = list(DEFAULT_EXCLUSIONS)
    _ensure_materials(state)


def _ensure_materials(state: BidFormState) -> None:
    """Initialize default materials sections if not set."""
    if not state.materials_sections:
        for section_name, items in DEFAULT_MATERIALS.items():
            state.materials_sections[section_name] = [
                MaterialItem(name=name, highlight=hl) for name, hl in items
            ]
    if not state.materials_section_order:
        state.materials_section_order = list(DEFAULT_MATERIALS_SECTION_ORDER)


@router.get("/spec", response_class=HTMLResponse)
async def spec_form_page(request: Request):
    """Render the spec form page with section totals and spec line items."""
    state = get_current_state()

    if not state:
        return RedirectResponse(url="/", status_code=302)

    _ensure_spec_items(state)

    # Build section totals in raw (Excel) order to match Bid Form
    totals_by_name = {st.section_name: st for st in state.section_totals}
    raw_sections = state.get_raw_sections()
    ordered_section_totals = [totals_by_name[s] for s in raw_sections if s in totals_by_name]

    # Collect excluded spec items for Add Alts section
    add_alts = []
    for section_name in raw_sections:
        for spec in state.spec_items.get(section_name, []):
            if spec.excluded:
                add_alts.append({"name": spec.name, "section": section_name, "price": spec.price})

    context = get_template_context(
        request,
        page="spec",
        bid_state=state,
        ordered_section_totals=ordered_section_totals,
        add_alts=add_alts,
        materials_sections=state.materials_sections,
        materials_section_order=state.materials_section_order,
        materials_brand=state.materials_brand,
    )

    return templates.TemplateResponse("spec.html", context)


@router.post("/spec/item/add", response_class=HTMLResponse)
async def add_spec_item(
    request: Request,
    section: str = Form(...),
    name: str = Form(...),
):
    """Add a spec item to a section and return the updated section body."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    trimmed = name.strip()
    if not trimmed:
        raise HTTPException(status_code=400, detail="Spec item name is required")

    state.spec_items.setdefault(section, []).append(SpecItem(name=trimmed))

    return _render_spec_section_body(request, state, section)


@router.post("/spec/item/delete", response_class=HTMLResponse)
async def delete_spec_item(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
):
    """Delete a spec item by index from a section and return the updated section body."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    items = state.spec_items.get(section, [])
    was_excluded = 0 <= index < len(items) and items[index].excluded
    if 0 <= index < len(items):
        items.pop(index)

    response = _render_spec_section_body(request, state, section)
    if was_excluded:
        response.headers["HX-Trigger"] = "add-alts-updated"
    return response


@router.post("/spec/item/name", response_class=HTMLResponse)
async def update_spec_item_name(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
    name: str = Form(...),
):
    """Update the name of a spec item."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    items = state.spec_items.get(section, [])
    if 0 <= index < len(items):
        items[index].name = name.strip()

    response = _render_spec_section_body(request, state, section)
    response.headers["HX-Trigger"] = "add-alts-updated"
    return response


@router.post("/spec/item/exclude", response_class=HTMLResponse)
async def toggle_spec_item_exclude(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
):
    """Toggle excluded state for a spec item and return the updated section body."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    items = state.spec_items.get(section, [])
    if 0 <= index < len(items):
        items[index].excluded = not items[index].excluded

    response = _render_spec_section_body(request, state, section)
    response.headers["HX-Trigger"] = "add-alts-updated"
    return response


@router.post("/spec/item/price", response_class=HTMLResponse)
async def update_spec_item_price(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
    price: str = Form(default=""),
):
    """Update the price for a spec item."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    items = state.spec_items.get(section, [])
    if 0 <= index < len(items):
        try:
            cleaned = price.replace(",", "").replace("$", "").strip()
            items[index].price = float(cleaned) if cleaned else None
        except ValueError:
            pass

    response = _render_spec_section_body(request, state, section)
    response.headers["HX-Trigger"] = "add-alts-updated"
    return response


@router.get("/spec/add-alts", response_class=HTMLResponse)
async def get_spec_add_alts(request: Request):
    """Return the Add Alts panel content for HTMX refresh."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    # Collect all excluded spec items across all sections, preserving raw order
    add_alts = []
    for section_name in state.get_raw_sections():
        for spec in state.spec_items.get(section_name, []):
            if spec.excluded:
                add_alts.append({"name": spec.name, "section": section_name, "price": spec.price})

    context = get_template_context(request, add_alts=add_alts)
    html = templates.get_template("partials/spec_add_alts.html").render(context)
    return HTMLResponse(html)


@router.post("/spec/section/label", response_class=HTMLResponse)
async def update_spec_section_label(
    request: Request,
    section: str = Form(...),
    label: str = Form(default=""),
):
    """Update the label for a spec section and return the updated header."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    trimmed = label.strip()
    if trimmed:
        state.spec_section_labels[section] = trimmed
    else:
        state.spec_section_labels.pop(section, None)

    # Find the matching section total
    section_total = None
    for st in state.section_totals:
        if st.section_name == section:
            section_total = st
            break

    if not section_total:
        raise HTTPException(status_code=404, detail="Section not found")

    context = get_template_context(
        request,
        section_name=section,
        section_total=section_total,
        section_label=trimmed,
        section_id=section.replace(" ", "-").replace("/", "-"),
    )
    html = templates.get_template("partials/spec_section_header.html").render(context)
    return HTMLResponse(html)


def _render_spec_section_body(request: Request, state: BidFormState, section: str) -> HTMLResponse:
    """Render the spec item rows for a single section."""
    items = state.spec_items.get(section, [])
    context = get_template_context(
        request,
        section_name=section,
        section_spec_items=items,
    )
    html = templates.get_template("partials/spec_section_body.html").render(context)
    return HTMLResponse(html)


@router.post("/spec/exclusion/add", response_class=HTMLResponse)
async def add_spec_exclusion(
    request: Request,
    name: str = Form(...),
):
    """Add a normal exclusion."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    trimmed = name.strip()
    if not trimmed:
        raise HTTPException(status_code=400, detail="Exclusion name is required")
    state.spec_exclusions.append(trimmed)
    return _render_exclusions_body(request, state)


@router.post("/spec/exclusion/delete", response_class=HTMLResponse)
async def delete_spec_exclusion(
    request: Request,
    index: int = Form(...),
):
    """Delete a normal exclusion by index."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    if 0 <= index < len(state.spec_exclusions):
        state.spec_exclusions.pop(index)
    return _render_exclusions_body(request, state)


def _render_exclusions_body(request: Request, state: BidFormState) -> HTMLResponse:
    """Render the exclusions list."""
    context = get_template_context(
        request,
        exclusions=state.spec_exclusions,
    )
    html = templates.get_template("partials/spec_exclusions_body.html").render(context)
    return HTMLResponse(html)


def _render_materials_body(request: Request, state: BidFormState) -> HTMLResponse:
    """Render the materials sections."""
    context = get_template_context(
        request,
        materials_sections=state.materials_sections,
        materials_section_order=state.materials_section_order,
        materials_brand=state.materials_brand,
    )
    html = templates.get_template("partials/spec_materials_body.html").render(context)
    return HTMLResponse(html)


@router.post("/spec/material/add", response_class=HTMLResponse)
async def add_material_item(
    request: Request,
    section: str = Form(...),
    name: str = Form(...),
):
    """Add a material item to a section."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    trimmed = name.strip()
    if not trimmed:
        raise HTTPException(status_code=400, detail="Material item name is required")
    state.materials_sections.setdefault(section, []).append(MaterialItem(name=trimmed))
    return _render_materials_body(request, state)


@router.post("/spec/material/delete", response_class=HTMLResponse)
async def delete_material_item(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
):
    """Delete a material item from a section."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    items = state.materials_sections.get(section, [])
    if 0 <= index < len(items):
        items.pop(index)
    return _render_materials_body(request, state)


@router.post("/spec/material/highlight", response_class=HTMLResponse)
async def toggle_material_highlight(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
):
    """Toggle yellow highlight on a material item."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    items = state.materials_sections.get(section, [])
    if 0 <= index < len(items):
        items[index].highlight = not items[index].highlight
    return _render_materials_body(request, state)


@router.post("/spec/material/value")
async def update_material_value(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
    value: str = Form(default=""),
):
    """Update the value/description text for a material item (save only, no re-render)."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    items = state.materials_sections.get(section, [])
    if 0 <= index < len(items):
        items[index].value = value.strip()
    return Response(status_code=204)


@router.post("/spec/material/name")
async def update_material_name(
    request: Request,
    section: str = Form(...),
    index: int = Form(...),
    name: str = Form(...),
):
    """Update the name of a material item (save only, no re-render)."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    trimmed = name.strip()
    if not trimmed:
        return Response(status_code=204)
    items = state.materials_sections.get(section, [])
    if 0 <= index < len(items):
        items[index].name = trimmed
    return Response(status_code=204)


@router.post("/spec/material/section/add", response_class=HTMLResponse)
async def add_material_section(
    request: Request,
    name: str = Form(...),
):
    """Add a new materials section."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    trimmed = name.strip()
    if not trimmed:
        raise HTTPException(status_code=400, detail="Section name is required")
    if trimmed not in state.materials_sections:
        state.materials_sections[trimmed] = []
        state.materials_section_order.append(trimmed)
    return _render_materials_body(request, state)


@router.post("/spec/material/section/delete", response_class=HTMLResponse)
async def delete_material_section(
    request: Request,
    section: str = Form(...),
):
    """Delete a materials section and all its items."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")
    _ensure_spec_items(state)
    state.materials_sections.pop(section, None)
    if section in state.materials_section_order:
        state.materials_section_order.remove(section)
    return _render_materials_body(request, state)


@router.get("/spec/print", response_class=HTMLResponse)
async def spec_print_page(request: Request):
    """Render the print-friendly spec form page."""
    state = get_current_state()
    if not state:
        return RedirectResponse(url="/", status_code=302)

    _ensure_spec_items(state)

    # Build section totals in raw (Excel) order
    totals_by_name = {st.section_name: st for st in state.section_totals}
    raw_sections = state.get_raw_sections()
    ordered_section_totals = [totals_by_name[s] for s in raw_sections if s in totals_by_name]

    # Collect excluded spec items for Add Alts
    add_alts = []
    for section_name in raw_sections:
        for spec in state.spec_items.get(section_name, []):
            if spec.excluded:
                add_alts.append({"name": spec.name, "section": section_name, "price": spec.price})

    # Build project address line
    info = state.project_info
    address_parts = [p for p in [info.address, info.city] if p]
    project_address = ", ".join(address_parts) if address_parts else state.project_name

    context = get_template_context(
        request,
        page="spec",
        bid_state=state,
        ordered_section_totals=ordered_section_totals,
        add_alts=add_alts,
        project_address=project_address,
        today_date=datetime.now().strftime("%Y-%m-%d"),
        materials_sections=state.materials_sections,
        materials_section_order=state.materials_section_order,
        materials_brand=state.materials_brand,
    )

    return templates.TemplateResponse("spec_print.html", context)


@router.get("/spec/export")
async def export_spec(request: Request, format: str = "xlsx"):
    """Export spec form as XLSX or PDF."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    _ensure_spec_items(state)

    filename_base = state.project_name or "spec"

    if format == "xlsx":
        content = _export_spec_xlsx(state)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}-spec.xlsx"'},
        )
    elif format == "pdf":
        content = _export_spec_pdf(state)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}-spec.pdf"'},
        )
    else:
        raise HTTPException(status_code=400, detail="Unsupported format. Use xlsx or pdf.")


def _export_spec_xlsx(state: BidFormState) -> bytes:
    """Build an XLSX workbook for the spec form matching print layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Spec Form"

    # Fonts
    font_title = Font(name="Cambria", bold=True, size=11, underline="single")
    font_section = Font(name="Cambria", bold=True, size=12)
    font_normal = Font(name="Cambria", size=11)
    font_exclude = Font(name="Cambria", italic=True, size=10, color="FF0000")
    font_pricing_hdr = Font(name="Cambria", bold=True, size=14)
    font_pricing = Font(name="Cambria", size=11)
    font_pricing_bold = Font(name="Cambria", bold=True, size=11)
    font_net_wrap = Font(name="Cambria", italic=True, size=11, color="FF0000")
    font_alt_hdr = Font(name="Cambria", bold=True, size=12)
    currency_fmt = '"$"#,##0.00'
    thin_top = Border(top=Side(style="thin"))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    info = state.project_info
    today = datetime.now().strftime("%B %d, %Y")

    font_hdr_lbl = Font(name="Cambria", bold=True, size=10)
    font_hdr_val = Font(name="Cambria", size=10)
    box_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    row = 1

    # ── Header Table 1: Developer/Address/City | Date/Contact/Phone ──
    hdr1_rows = [
        ("DEVELOPER:", info.developer or "", "DATE:", today),
        ("ADDRESS:", info.address or "", "CONTACT:", info.contact or ""),
        ("CITY:", info.city or "", "PHONE:", info.phone or ""),
    ]
    for ll, lv, rl, rv in hdr1_rows:
        for col, val, fnt in [(1, ll, font_hdr_lbl), (2, lv, font_hdr_val),
                               (3, rl, font_hdr_lbl), (4, rv, font_hdr_val)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt
            c.border = box_border
        row += 1

    row += 1  # blank row

    # ── Header Table 2: Project/Units/City | Plans dated ──
    hdr2_rows = [
        ("PROJECT", state.project_name or "", "PLANS", "DATED"),
        ("UNITS", info.units_text or "", "ARCHITECTURAL", _fmt_date(info.arch_date)),
        ("CITY", info.project_city or "", "LANDSCAPE", _fmt_date(info.landscape_date) or "Excluded"),
        ("", "", "INTERIOR DESIGN", _fmt_date(info.interior_design_date)),
        ("", "", "OWNER SPECS", _fmt_date(info.owner_specs_date) or "NA"),
    ]
    for ll, lv, rl, rv in hdr2_rows:
        for col, val, fnt in [(1, ll, font_hdr_lbl), (2, lv, font_hdr_val),
                               (3, rl, font_hdr_lbl), (4, rv, font_hdr_val)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt
            c.border = box_border
        row += 1

    row += 2  # blank rows before scope

    # Use raw section order
    totals_by_name = {st.section_name: st for st in state.section_totals}
    raw_sections = state.get_raw_sections()

    # Collect add alts while iterating
    add_alts = []

    # -- SCOPE SECTIONS --
    for section_name in raw_sections:
        items = state.spec_items.get(section_name, [])
        included = [s for s in items if not s.excluded]
        excluded = [s for s in items if s.excluded]

        if not included and not excluded:
            continue

        for s in excluded:
            add_alts.append({"name": s.name, "section": section_name, "price": s.price})

        # Section title
        ws.cell(row=row, column=1, value=section_name.upper()).font = font_section
        row += 1

        # Included items
        for spec in included:
            ws.cell(row=row, column=1, value=spec.name).font = font_normal
            row += 1

        # Excluded items
        for spec in excluded:
            ws.cell(row=row, column=1, value=f"Excludes {spec.name}").font = font_exclude
            row += 1

        row += 1  # blank row between sections

    # -- PRICING TABLE --
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    pricing_cell = ws.cell(row=row, column=1, value="PRICING")
    pricing_cell.font = font_pricing_hdr
    pricing_cell.alignment = Alignment(horizontal="center")
    row += 1

    # Amount header
    ws.cell(row=row, column=4, value="Amount").font = font_pricing_bold
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    row += 1

    # Section rows
    for section_name in raw_sections:
        st = totals_by_name.get(section_name)
        if not st or st.total <= 0:
            continue
        ws.cell(row=row, column=1, value=section_name).font = font_pricing
        total_cell = ws.cell(row=row, column=4, value=st.total)
        total_cell.font = font_pricing
        total_cell.number_format = currency_fmt
        total_cell.alignment = Alignment(horizontal="right")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = font_pricing_bold
    grand_cell = ws.cell(row=row, column=4, value=state.grand_total)
    grand_cell.font = font_pricing_bold
    grand_cell.number_format = currency_fmt
    grand_cell.alignment = Alignment(horizontal="right")
    grand_cell.border = thin_top
    ws.cell(row=row, column=1).border = thin_top
    row += 2

    # Net wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    nw_cell = ws.cell(row=row, column=1, value="Pricing Net Wrap Liability Insurance")
    nw_cell.font = font_net_wrap
    nw_cell.alignment = Alignment(horizontal="center")
    row += 2

    # -- NORMAL EXCLUSIONS --
    if state.spec_exclusions:
        font_excl_hdr = Font(name="Cambria", bold=True, size=12, color="FF0000")
        font_excl_item = Font(name="Cambria", size=11)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        excl_cell = ws.cell(row=row, column=1, value="EXCLUSIONS")
        excl_cell.font = font_excl_hdr
        excl_cell.alignment = Alignment(horizontal="center")
        row += 1
        for i in range(0, len(state.spec_exclusions), 2):
            ws.cell(row=row, column=1, value=state.spec_exclusions[i]).font = font_excl_item
            if i + 1 < len(state.spec_exclusions):
                ws.cell(row=row, column=3, value=state.spec_exclusions[i + 1]).font = font_excl_item
            row += 1
        row += 1

    # -- ADD ALTERNATES --
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    if add_alts:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        alt_cell = ws.cell(row=row, column=1, value="ADD ALTERNATES")
        alt_cell.font = font_alt_hdr
        alt_cell.alignment = Alignment(horizontal="center")
        alt_cell.border = Border(bottom=Side(style="thin"))
        row += 1

        for alt in add_alts:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            name_cell = ws.cell(row=row, column=1, value=f"{alt['name']} at {alt['section']}")
            name_cell.font = font_normal
            name_cell.border = thin_border
            ws.cell(row=row, column=2).border = thin_border
            dollar_cell = ws.cell(row=row, column=3, value="$")
            dollar_cell.font = font_normal
            dollar_cell.border = thin_border
            price = alt.get("price")
            amt_cell = ws.cell(row=row, column=4, value=price if price else None)
            amt_cell.font = font_normal
            amt_cell.border = thin_border
            amt_cell.alignment = Alignment(horizontal="right")
            if price:
                amt_cell.number_format = '#,##0.00'
            row += 1

    # -- ADDITIONAL WORK RATES --
    font_bold_normal = Font(name="Cambria", size=11, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="ADDITIONAL WORK CHARGED AT:").font = font_bold_normal
    ws.cell(row=row, column=4, value="$73.00/HR").font = font_bold_normal
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="1/2 Time OT Work").font = font_normal
    ws.cell(row=row, column=4, value="$37.00/HR").font = font_normal
    row += 2

    # -- MATERIALS INCLUDED IN BID --
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if state.materials_section_order:
        ws.cell(row=row, column=1, value=f"MATERIALS INCLUDED IN BID: {state.materials_brand}").font = font_bold_normal
        row += 1

        for section_label in state.materials_section_order:
            mat_items = state.materials_sections.get(section_label, [])
            ws.cell(row=row, column=1, value=section_label).font = font_bold_normal
            row += 1
            for item in mat_items:
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = font_normal
                ws.cell(row=row, column=1, value=item.name)
                if item.value:
                    ws.cell(row=row, column=2, value=item.value)
                row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_spec_pdf(state: BidFormState) -> bytes:
    """Build a professional proposal PDF matching print layout."""
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="Letter")
    pdf.set_auto_page_break(auto=True, margin=22)
    pdf.set_margins(left=18, top=16, right=18)
    pdf.add_page()

    page_w = pdf.w - pdf.l_margin - pdf.r_margin
    # Inset: content area inside margins with internal padding
    pad = 3  # mm inner padding on each side
    content_w = page_w - 2 * pad
    content_x = pdf.l_margin + pad

    info = state.project_info
    today = datetime.now().strftime("%B %d, %Y")
    lw = 0.3  # default line width

    # ── Helper: draw a structured line ──
    def hline(y, x1=None, x2=None, width=0.3):
        pdf.set_line_width(width)
        pdf.line(x1 or content_x, y, x2 or (content_x + content_w), y)
        pdf.set_line_width(lw)

    # ────────────────────────────────────────────────────
    # 1. HEADER TABLE  (Developer/Address/City | Date/Contact/Phone)
    # ────────────────────────────────────────────────────
    half_w = content_w / 2
    lbl_w = 22
    val_w = half_w - lbl_w
    row_h = 7
    header_rows = [
        ("DEVELOPER:", info.developer or "", "DATE:", today),
        ("ADDRESS:", info.address or "", "CONTACT:", info.contact or ""),
        ("CITY:", info.city or "", "PHONE:", info.phone or ""),
    ]
    start_y = pdf.get_y()
    x0 = content_x
    for i, (ll, lv, rl, rv) in enumerate(header_rows):
        y = start_y + i * row_h
        pdf.set_xy(x0, y)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(lbl_w, row_h, ll)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(val_w, row_h, lv)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(lbl_w, row_h, rl)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(val_w, row_h, rv)

    total_h = len(header_rows) * row_h
    # Outer box (0.75pt)
    pdf.set_line_width(0.6)
    pdf.rect(x0, start_y, content_w, total_h)
    # Vertical divider between halves
    pdf.line(x0 + half_w, start_y, x0 + half_w, start_y + total_h)
    pdf.set_line_width(lw)
    # Light horizontal lines between rows
    pdf.set_draw_color(180, 180, 180)
    for i in range(1, len(header_rows)):
        y = start_y + i * row_h
        pdf.line(x0, y, x0 + content_w, y)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_y(start_y + total_h + 4)

    # ────────────────────────────────────────────────────
    # 2. PROJECT INFO TABLE  (Project/Units/City | Plans dated)
    # ────────────────────────────────────────────────────
    left_lbl = 17
    left_val = half_w - left_lbl
    right_lbl = 32
    right_val = half_w - right_lbl
    proj_rows = [
        ("PROJECT", state.project_name or "", "PLANS", "DATED"),
        ("UNITS", info.units_text or "", "ARCHITECTURAL", _fmt_date(info.arch_date)),
        ("CITY", info.project_city or "", "LANDSCAPE", _fmt_date(info.landscape_date) or "Excluded"),
        ("", "", "INTERIOR DESIGN", _fmt_date(info.interior_design_date)),
        ("", "", "OWNER SPECS", _fmt_date(info.owner_specs_date) or "NA"),
    ]
    start_y = pdf.get_y()
    for i, (ll, lv, rl, rv) in enumerate(proj_rows):
        y = start_y + i * row_h
        pdf.set_xy(x0, y)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(left_lbl, row_h, ll)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(left_val, row_h, lv)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(right_lbl, row_h, rl)
        # "DATED" header is bold centered; landscape "Excluded" is red
        if i == 0:
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(right_val, row_h, rv, align="C")
        elif rl == "LANDSCAPE" and rv.lower() == "excluded":
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(204, 0, 0)
            pdf.cell(right_val, row_h, rv, align="C")
            pdf.set_text_color(0, 0, 0)
        else:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(right_val, row_h, rv, align="C")

    total_h = len(proj_rows) * row_h
    pdf.set_line_width(0.6)
    pdf.rect(x0, start_y, content_w, total_h)
    pdf.line(x0 + half_w, start_y, x0 + half_w, start_y + total_h)
    pdf.set_line_width(lw)
    pdf.set_draw_color(180, 180, 180)
    for i in range(1, len(proj_rows)):
        y = start_y + i * row_h
        pdf.line(x0, y, x0 + content_w, y)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_y(start_y + total_h + 6)

    # ────────────────────────────────────────────────────
    # 3. SCOPE SECTIONS  (from spec items)
    # ────────────────────────────────────────────────────
    totals_by_name = {st.section_name: st for st in state.section_totals}
    raw_sections = state.get_raw_sections()
    add_alts = []

    for section_name in raw_sections:
        items = state.spec_items.get(section_name, [])
        included = [s for s in items if not s.excluded]
        excluded = [s for s in items if s.excluded]
        if not included and not excluded:
            continue
        for s in excluded:
            add_alts.append({"name": s.name, "section": section_name, "price": s.price})

        # Page break check: title + first 3 items kept together
        needed = 10 + min(len(included), 3) * 6 + 6
        if pdf.get_y() + needed > pdf.h - 25:
            pdf.add_page()

        # 12pt space before section
        pdf.ln(4)

        # Section title: uppercase, bold, 11pt — with label if set
        section_label = state.spec_section_labels.get(section_name, "")
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_x(content_x)
        if section_label:
            pdf.cell(content_w * 0.6, 7, section_name.upper())
            # Yellow-highlighted label
            pdf.set_fill_color(255, 255, 0)
            pdf.set_font("Helvetica", "B", 10)
            label_w = pdf.get_string_width(section_label) + 6
            pdf.set_x(content_x + content_w - label_w)
            pdf.cell(label_w, 7, section_label, fill=True,
                     new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.cell(content_w, 7, section_name.upper(), new_x="LMARGIN", new_y="NEXT")
        # Thin rule across full width
        hline(pdf.get_y(), width=0.5)
        pdf.ln(2)

        # Included items — indented
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(0, 0, 0)
        for spec in included:
            pdf.set_x(content_x + 4)
            pdf.cell(content_w - 4, 5.5, spec.name, new_x="LMARGIN", new_y="NEXT")

        # Excluded items — dark red italic
        if excluded:
            pdf.set_font("Helvetica", "I", 10)
            pdf.set_text_color(170, 0, 0)
            for spec in excluded:
                pdf.set_x(content_x + 4)
                pdf.cell(content_w - 4, 5.5, f"Excluded {spec.name}",
                         new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)

        pdf.ln(1)

    # ────────────────────────────────────────────────────
    # 4. PRICING TABLE  (boxed, structured)
    # ────────────────────────────────────────────────────
    pdf.ln(4)
    # Collect visible sections for sizing
    visible_sections = [(s, totals_by_name[s]) for s in raw_sections
                        if s in totals_by_name and totals_by_name[s].total > 0]
    box_h = 10 + 7 + len(visible_sections) * 7 + 7 + 10  # header + col hdr + rows + total + net
    if pdf.get_y() + box_h > pdf.h - 25:
        pdf.add_page()

    col_name_w = 55
    col_amt_w = 35
    table_w = col_name_w + col_amt_w
    table_x = content_x + (content_w - table_w) / 2

    # Pricing box start
    box_y = pdf.get_y()

    # "PRICING" header bar (gray bg)
    pdf.set_fill_color(245, 245, 245)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(content_x)
    pdf.cell(content_w, 10, "PRICING", align="C", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    hdr_bottom = pdf.get_y()

    # Column header
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_x(table_x)
    pdf.cell(col_name_w, 7, "")
    pdf.cell(col_amt_w, 7, "Amount", align="C", new_x="LMARGIN", new_y="NEXT")
    hline(pdf.get_y(), table_x, table_x + table_w, 0.3)

    # Section rows
    pdf.set_font("Helvetica", "", 10)
    for section_name, st in visible_sections:
        pdf.set_x(table_x)
        pdf.cell(col_name_w, 7, section_name)
        pdf.cell(col_amt_w, 7, f"${st.total:,.2f}", align="R",
                 new_x="LMARGIN", new_y="NEXT")

    # Total row
    hline(pdf.get_y(), table_x, table_x + table_w, 0.6)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_x(table_x)
    pdf.cell(col_name_w, 7, "Total")
    pdf.cell(col_amt_w, 7, f"${state.grand_total:,.2f}", align="R",
             new_x="LMARGIN", new_y="NEXT")

    # Net wrap
    pdf.ln(1)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(204, 0, 0)
    pdf.set_x(content_x)
    pdf.cell(content_w, 6, "Pricing Net Wrap Liability Insurance",
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_text_color(0, 0, 0)

    # Outer box around entire pricing block
    box_bottom = pdf.get_y() + 2
    pdf.set_line_width(0.6)
    pdf.rect(content_x, box_y, content_w, box_bottom - box_y)
    # Line below header bar
    hline(hdr_bottom, width=0.6)
    pdf.set_line_width(lw)
    pdf.set_y(box_bottom + 4)

    # ────────────────────────────────────────────────────
    # 5. EXCLUSIONS  (boxed, two-column)
    # ────────────────────────────────────────────────────
    if state.spec_exclusions:
        excl = state.spec_exclusions
        rows_needed = (len(excl) + 1) // 2
        needed_h = 12 + rows_needed * 5.5 + 4
        if pdf.get_y() + needed_h > pdf.h - 25:
            pdf.add_page()

        box_y = pdf.get_y()

        # Title
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(204, 0, 0)
        pdf.set_x(content_x)
        pdf.cell(content_w, 8, "EXCLUSIONS", new_x="LMARGIN", new_y="NEXT", align="C")
        hline(pdf.get_y(), width=0.3)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)

        # Two-column items
        pdf.set_font("Helvetica", "", 9)
        col_w = content_w / 2
        for i in range(0, len(excl), 2):
            pdf.set_x(content_x + 3)
            pdf.cell(col_w - 3, 5, excl[i])
            if i + 1 < len(excl):
                pdf.cell(col_w, 5, excl[i + 1],
                         new_x="LMARGIN", new_y="NEXT")
            else:
                pdf.cell(col_w, 5, "", new_x="LMARGIN", new_y="NEXT")

        box_bottom = pdf.get_y() + 2
        pdf.set_line_width(0.6)
        pdf.rect(content_x, box_y, content_w, box_bottom - box_y)
        pdf.set_line_width(lw)
        pdf.set_y(box_bottom + 4)

    # ────────────────────────────────────────────────────
    # 6. ADDITIONAL WORK RATES  (clean mini-table, no box)
    # ────────────────────────────────────────────────────
    if pdf.get_y() + 20 > pdf.h - 25:
        pdf.add_page()

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_x(content_x)
    pdf.cell(content_w * 0.65, 7, "ADDITIONAL WORK CHARGED AT:")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(content_w * 0.35, 7, "$73.00/HR", align="R",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_x(content_x)
    pdf.cell(content_w * 0.65, 7, "1/2 Time OT Work")
    pdf.cell(content_w * 0.35, 7, "$37.00/HR", align="R",
             new_x="LMARGIN", new_y="NEXT")

    # ────────────────────────────────────────────────────
    # 7. ADD ALTERNATES
    # ────────────────────────────────────────────────────
    if add_alts:
        pdf.ln(4)
        row_h = 6
        if pdf.get_y() + 12 + len(add_alts) * row_h > pdf.h - 25:
            pdf.add_page()

        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_x(content_x)
        pdf.cell(content_w, 8, "ADD ALTERNATES", new_x="LMARGIN", new_y="NEXT", align="C")
        hline(pdf.get_y(), width=0.5)

        # Bordered table: name | $ | amount
        name_w = content_w - 12 - 30  # remaining space for name
        dollar_w = 12
        amt_w = 30

        pdf.set_font("Helvetica", "", 10)
        for alt in add_alts:
            y = pdf.get_y()
            pdf.set_x(content_x)
            section_tag = f" at {alt['section']}"
            pdf.cell(name_w, row_h, alt["name"] + section_tag)
            pdf.cell(dollar_w, row_h, "$")
            price = alt.get("price")
            pdf.cell(amt_w, row_h, f"{price:,.2f}" if price else "",
                     align="R", new_x="LMARGIN", new_y="NEXT")
            # Draw cell borders
            pdf.set_line_width(0.3)
            pdf.rect(content_x, y, name_w, row_h)
            pdf.rect(content_x + name_w, y, dollar_w, row_h)
            pdf.rect(content_x + name_w + dollar_w, y, amt_w, row_h)
        pdf.set_line_width(lw)

    # ────────────────────────────────────────────────────
    # 8. MATERIALS INCLUDED IN BID  (structured subsections)
    # ────────────────────────────────────────────────────
    if state.materials_section_order:
        pdf.ln(6)
        if pdf.get_y() + 50 > pdf.h - 25:
            pdf.add_page()

        # Main title
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_x(content_x)
        pdf.cell(content_w, 7, f"MATERIALS INCLUDED IN BID: {state.materials_brand}",
                 new_x="LMARGIN", new_y="NEXT")

        mat_col1 = 55
        mat_col2 = content_w - mat_col1
        mat_row_h = 6

        for section_label in state.materials_section_order:
            mat_items = state.materials_sections.get(section_label, [])
            # Subsection title row — full width, light gray bg
            y = pdf.get_y()
            pdf.set_fill_color(245, 245, 245)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_x(content_x)
            pdf.cell(content_w, mat_row_h, section_label, fill=True,
                     new_x="LMARGIN", new_y="NEXT")
            pdf.rect(content_x, y, content_w, mat_row_h)
            # Item rows — bordered cells
            pdf.set_font("Helvetica", "", 10)
            for item in mat_items:
                y = pdf.get_y()
                pdf.set_x(content_x)
                pdf.cell(mat_col1, mat_row_h, f"  {item.name}")
                pdf.cell(mat_col2, mat_row_h, item.value or "",
                         new_x="LMARGIN", new_y="NEXT")
                pdf.rect(content_x, y, mat_col1, mat_row_h)
                pdf.rect(content_x + mat_col1, y, mat_col2, mat_row_h)

    # ────────────────────────────────────────────────────
    # 9. SIGNATURE SECTION
    # ────────────────────────────────────────────────────
    if pdf.get_y() + 45 > pdf.h - 25:
        pdf.add_page()

    pdf.ln(10)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.set_x(content_x)
    pdf.cell(content_w, 5, "This proposal is valid for 30 days from the date above.",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    sig_w = content_w * 0.44
    gap_w = content_w * 0.12

    # Left signature block
    sig_y = pdf.get_y()
    pdf.set_font("Helvetica", "", 9)
    pdf.set_x(content_x)
    pdf.cell(sig_w, 5, "Authorized Signature - RCW Painting",
             new_x="LMARGIN", new_y="NEXT")
    line_y = pdf.get_y() + 10
    pdf.line(content_x, line_y, content_x + sig_w, line_y)
    pdf.set_y(line_y + 2)
    pdf.set_x(content_x)
    pdf.cell(sig_w, 5, "Date: _____________", new_x="LMARGIN", new_y="NEXT")

    # Right signature block
    right_x = content_x + sig_w + gap_w
    pdf.set_xy(right_x, sig_y)
    pdf.cell(sig_w, 5, "Acceptance - Customer",
             new_x="LMARGIN", new_y="NEXT")
    pdf.line(right_x, line_y, right_x + sig_w, line_y)
    pdf.set_xy(right_x, line_y + 2)
    pdf.cell(sig_w, 5, "Date: _____________", new_x="LMARGIN", new_y="NEXT")

    pdf.set_y(line_y + 10)

    # ────────────────────────────────────────────────────
    # 10. FOOTER
    # ────────────────────────────────────────────────────
    pdf.set_draw_color(180, 180, 180)
    hline(pdf.get_y(), width=0.3)
    pdf.set_draw_color(0, 0, 0)
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(140, 140, 140)
    pdf.set_x(content_x)
    pdf.cell(content_w, 5, f"RCW Painting  |  {today}", align="C",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    buf = BytesIO()
    pdf.output(buf)
    return buf.getvalue()


@router.get("/logic", response_class=HTMLResponse)
async def logic_page(request: Request):
    """Render the logic/business rules page."""
    state = get_current_state()

    if not state:
        # No bid form available, create a sample one for testing
        from app.ui.excel_mapper import create_sample_bid_form
        sample_state = create_sample_bid_form()
        set_state("sample", sample_state)
        state = sample_state

    # Build sections dictionary with items grouped by section
    sections_dict = {}
    for section_name in state.get_raw_sections():
        sections_dict[section_name] = state.get_raw_items_by_section(section_name)

    # Build concrete worked examples so logic is easy to validate.
    logic_examples = []
    for item in [i for i in state.raw_items if i.qty > 0 and not i.excluded][:5]:
        difficulty_add = item.difficulty_adders.get(item.difficulty, 0.0)
        base_plus_difficulty = item.unit_price_base + difficulty_add
        effective_unit = base_plus_difficulty * item.mult
        row_total = item.qty * effective_unit

        logic_examples.append({
            "name": item.name,
            "section": item.section,
            "uom": item.uom,
            "qty": item.qty,
            "base_price": item.unit_price_base,
            "difficulty_level": item.difficulty,
            "difficulty_add": difficulty_add,
            "manual_multiplier": item.mult,
            "effective_unit": effective_unit,
            "row_total": row_total,
        })

    # Compute summary stats for the instructions page
    active_items = [i for i in state.raw_items if not i.excluded]
    priced_items = [i for i in active_items if i.qty > 0]
    excluded_count = sum(1 for i in state.raw_items if i.excluded)
    exclusion_count = sum(1 for i in state.raw_items if i.is_exclusion)
    alternate_count = sum(1 for i in state.raw_items if i.is_alternate)
    sections_list = state.get_raw_sections()

    # Get logic rules and calculations for display
    context = get_template_context(
        request,
        page="logic",
        bid_state=state,
        sections=sections_dict,  # Now a dictionary of section_name -> list of items
        total_items=state.total_items,
        grand_total=state.grand_total,
        section_totals=state.section_totals,
        logic_examples=logic_examples,
        priced_item_count=len(priced_items),
        excluded_count=excluded_count,
        exclusion_count=exclusion_count,
        alternate_count=alternate_count,
        sections_list=sections_list,
    )

    return templates.TemplateResponse("logic.html", context)


@router.get("/print", response_class=HTMLResponse)
async def print_page(request: Request):
    """Render the print-friendly bid proposal page."""
    from datetime import datetime

    state = get_current_state()
    if not state:
        # Redirect to home if no bid data
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/", status_code=302)

    # Use raw sections (Excel order)
    sections = state.get_raw_sections()

    sections_data = {}
    for section_name in sections:
        sections_data[section_name] = state.get_raw_items_by_section(section_name)

    # Calculate unit count and total SF from raw items
    unit_count = 0
    total_sf = 0.0
    alternates = []

    for item in state.raw_items:
        if item.excluded:
            continue
        # Count units (look for unit count items)
        name_lower = item.name.lower()
        if 'unit' in name_lower and 'count' in name_lower:
            unit_count += int(item.qty)
        # Sum up SF items for total
        if item.uom == 'SF':
            total_sf += item.qty
        # Collect alternates
        if item.is_alternate:
            alternates.append(item)

    _ensure_materials(state)

    # Collect excluded spec items for Add Alts on proposal
    add_alts = []
    for section_name in sections:
        for spec in state.spec_items.get(section_name, []):
            if spec.excluded:
                add_alts.append({"name": spec.name, "section": section_name, "price": spec.price})

    context = get_template_context(
        request,
        page="print",
        bid_state=state,
        sections=sections,
        sections_data=sections_data,
        project_info=state.project_info,
        today_date=datetime.now().strftime("%B %d, %Y"),
        unit_count=unit_count,
        total_sf=total_sf,
        alternates=alternates,
        add_alts=add_alts,
        spec_items=state.spec_items,
        spec_section_labels=state.spec_section_labels,
        materials_sections=state.materials_sections,
        materials_section_order=state.materials_section_order,
        materials_brand=state.materials_brand,
    )

    return templates.TemplateResponse("print.html", context)


@router.post("/bid/project-info", response_class=HTMLResponse)
async def update_project_info(
    request: Request,
    developer: str = Form(default=""),
    address: str = Form(default=""),
    city: str = Form(default=""),
    contact: str = Form(default=""),
    phone: str = Form(default=""),
    email: str = Form(default=""),
    project_name: str = Form(default=""),
    project_city: str = Form(default=""),
    units_text: str = Form(default=""),
    plans_date: str = Form(default=""),
    arch_date: str = Form(default=""),
    landscape_date: str = Form(default=""),
    interior_design_date: str = Form(default=""),
    owner_specs_date: str = Form(default=""),
):
    """Update project info (saves silently, no re-render)."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    if project_name.strip():
        state.project_name = project_name.strip()

    state.project_info = ProjectInfo(
        developer=developer or None,
        address=address or None,
        city=city or None,
        contact=contact or None,
        phone=phone or None,
        email=email or None,
        project_city=project_city or None,
        units_text=units_text or None,
        plans_date=plans_date or None,
        arch_date=arch_date or None,
        landscape_date=landscape_date or None,
        interior_design_date=interior_design_date or None,
        owner_specs_date=owner_specs_date or None,
    )

    return HTMLResponse(status_code=204)


@router.get("/bid/project-info-panel", response_class=HTMLResponse)
async def get_project_info_panel(request: Request):
    """Return the project info header partial for HTMX refresh."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    context = get_template_context(
        request,
        bid_state=state,
        **_project_header_vars(state),
    )

    html = templates.get_template("partials/project_info_form.html").render(context)
    return HTMLResponse(html)


@router.get("/bid/export", response_class=HTMLResponse)
async def export_bid(request: Request, format: str = "json"):
    """Export the current bid in various formats."""
    state = get_current_state()
    if not state:
        raise HTTPException(status_code=404, detail="No active bid form")

    if format == "json":
        return state.model_dump_json(indent=2)
    elif format == "internal_xlsx":
        content = export_internal_bid_workbook(state)
        filename = f"{state.project_name or 'bid'}-internal.xlsx"
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    elif format == "proposal_xlsx":
        content = export_proposal_workbook(state)
        filename = f"{state.project_name or 'bid'}-proposal.xlsx"
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    elif format == "csv":
        # TODO: Implement CSV export
        raise HTTPException(status_code=501, detail="CSV export not yet implemented")
    else:
        raise HTTPException(status_code=400, detail="Unsupported export format")


@router.get("/bid/missing-stubs")
async def get_missing_catalog_stubs(request: Request):
    """
    Generate catalog stubs for missing extracted items.

    Returns JSON that can be merged into config/bid_catalog.json
    to add missing items and reduce warnings.
    """
    from app.ui.state import get_current_warnings

    # Get warnings from current session
    warnings = get_current_warnings()

    # Extract missing item IDs from warnings
    missing_ids = []
    for w in warnings:
        if isinstance(w, str) and "not found in catalog" in w:
            # Extract ID from warning like "Extracted item 'xxx.yyy' not found in catalog"
            start = w.find("'")
            end = w.find("'", start + 1)
            if start > 0 and end > start:
                missing_ids.append(w[start+1:end])

    if not missing_ids:
        return {"message": "No missing items found", "stubs": []}

    # Generate stubs
    stubs = []
    for item_id in missing_ids:
        parts = item_id.split('.')
        section_id = parts[0] if len(parts) > 1 else 'unknown'
        item_slug = parts[-1] if parts else 'unknown'
        label = item_slug.replace('_', ' ').title()

        stub = {
            "id": item_id,
            "label": label,
            "uom": "EA",  # Default - should be updated
            "rates": {"1": 0, "2": 0, "3": 0, "4": 0, "5": 0},
            "default_multiplier": 1.0
        }
        stubs.append(stub)

    return {
        "message": f"Found {len(stubs)} missing items. Add these to config/bid_catalog.json",
        "total_missing": len(stubs),
        "stubs": stubs
    }


@router.post("/bid/clear", response_class=RedirectResponse)
async def clear_bid(request: Request):
    """Clear the current bid and redirect to home."""
    from app.ui.state import clear_state
    clear_state()
    return RedirectResponse(url="/", status_code=303)
