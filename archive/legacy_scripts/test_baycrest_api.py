#!/usr/bin/env python3
"""
End-to-end test of Baycrest format through the API.
"""

import os
import sys
import json
import time
import tempfile
from pathlib import Path
import requests
import openpyxl
from openpyxl import Workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))


def create_baycrest_excel(file_path: str):
    """Create a Baycrest format Excel file for testing."""
    wb = Workbook()

    # Create "1 Bldg" sheet
    ws = wb.active
    ws.title = "1 Bldg"

    # Header row
    ws['A1'] = 'Section'
    ws['B1'] = 'Item'
    ws['C1'] = 'Main Qty'
    ws['D1'] = 'Secondary'
    ws['E1'] = 'Notes'

    # General section
    ws['A2'] = 'General'
    ws['B2'] = None
    ws['C2'] = None

    ws['B3'] = 'Unit Count'
    ws['C3'] = 150
    ws['D3'] = None
    ws['E3'] = 'Total units in building'

    ws['B4'] = 'Gross Building SF'
    ws['C4'] = 125000
    ws['D4'] = None

    # Corridors section
    ws['A5'] = 'Corridors'
    ws['B5'] = 'Flooring SF'
    ws['C5'] = 9500
    ws['E5'] = 'Corridor Notes'

    ws['B6'] = 'Wall SF'
    ws['C6'] = 15000

    ws['B7'] = 'Base LF'
    ws['C7'] = 1800

    # Exterior section
    ws['A8'] = 'Exterior'
    ws['B8'] = 'Stucco Wall SF'
    ws['C8'] = 30000

    ws['B9'] = 'Parapet LF'
    ws['C9'] = 500

    # Units section
    ws['A10'] = 'Units'
    ws['B10'] = 'Flooring SF'
    ws['C10'] = 95000

    ws['B11'] = 'Unit Doors'
    ws['C11'] = 600

    # Save the workbook
    wb.save(file_path)
    print(f"Created Baycrest Excel file: {file_path}")


def test_baycrest_api():
    """Test the Baycrest format through the API."""

    API_URL = "http://127.0.0.1:8006"
    API_KEY = "test-api-key"

    # Start a test server
    print("Starting test server on port 8006...")
    os.system("export DATABASE_URL='sqlite:///./data/test_baycrest_api.db' && "
              "export API_KEY='test-api-key' && "
              "export DISABLE_BOOTSTRAP_USERS=true && "
              "venv/bin/uvicorn app.main:app --port 8006 --host 127.0.0.1 &")

    # Wait for server to start
    time.sleep(3)

    try:
        # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_file = tmp.name

        # Create test Excel file
        create_baycrest_excel(test_file)

        print("\n" + "="*60)
        print("Testing Baycrest Format via API")
        print("="*60)

        # Upload file
        print("\n1. Uploading file with baycrest_v1 template...")
        with open(test_file, 'rb') as f:
            files = {'file': ('baycrest_test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {'template': 'baycrest_v1'}
            headers = {'X-API-Key': API_KEY}

            response = requests.post(f"{API_URL}/api/v1/jobs/", files=files, data=data, headers=headers)

        print(f"Upload response: {response.status_code}")
        assert response.status_code == 200, f"Upload failed: {response.text}"

        upload_result = response.json()
        job_id = upload_result['job_id']
        print(f"Job ID: {job_id}")

        # Poll for completion
        print("\n2. Polling for job completion...")
        max_attempts = 20
        for attempt in range(max_attempts):
            response = requests.get(f"{API_URL}/api/v1/jobs/{job_id}", headers={'X-API-Key': API_KEY})
            assert response.status_code == 200, f"Status check failed: {response.text}"

            job_status = response.json()
            print(f"  Attempt {attempt + 1}: Status = {job_status['status']}, Progress = {job_status['progress']}%")

            if job_status['status'] == 'SUCCEEDED':
                print("✓ Job completed successfully!")
                break

            if job_status['status'] == 'FAILED':
                print(f"✗ Job failed: {job_status.get('error', {}).get('message', 'Unknown error')}")
                assert False, "Job processing failed"

            time.sleep(1)
        else:
            assert False, "Job did not complete within timeout"

        # Check results
        print("\n3. Checking results...")
        result = job_status['result']
        qa = job_status['qa']

        # Check for raw_rows and raw_data (Baycrest specific)
        assert 'raw_rows' in result, "Result should contain raw_rows for Baycrest format"
        assert 'raw_data' in result, "Result should contain raw_data for Baycrest format"
        print(f"✓ raw_rows present: {len(result['raw_rows'])} rows")
        print(f"✓ raw_data present: {len(result['raw_data'])} rows")

        # Check sections
        assert 'sections' in result, "Result should contain sections"
        print(f"✓ Sections mapped: {len(result['sections'])}")
        for section in result['sections']:
            print(f"  - {section['name']}: {len(section['items'])} items")

        # Check QA stats
        assert 'stats' in qa, "QA should always contain stats (NextJS requirement)"
        print(f"✓ QA stats present:")
        for key, value in qa['stats'].items():
            print(f"    {key}: {value}")

        # Check specific values
        print("\n4. Verifying specific extractions...")

        # Check raw_rows
        unit_count_raw = next((row for row in result['raw_rows']
                               if row.get('B') == 'Unit Count'), None)
        assert unit_count_raw is not None, "Should find Unit Count in raw_rows"
        assert unit_count_raw['C'] == 150, "Unit Count should be 150 in raw_rows"
        print(f"✓ Unit Count in raw_rows: {unit_count_raw['C']}")

        # Check raw_data
        unit_count_data = next((item for item in result['raw_data']
                               if item['classification'] == 'Unit Count'), None)
        assert unit_count_data is not None, "Should find Unit Count in raw_data"
        assert unit_count_data['measures'][0]['value'] == 150, "Unit Count should be 150"
        print(f"✓ Unit Count in raw_data: {unit_count_data['measures'][0]['value']} {unit_count_data['measures'][0]['uom']}")

        # Check mapped sections
        general_section = next((s for s in result['sections'] if s['name'] == 'General'), None)
        assert general_section is not None, "Should have General section"
        unit_count_item = next((i for i in general_section['items'] if i['key'] == 'Unit Count'), None)
        assert unit_count_item is not None, "Should have Unit Count in General section"
        print(f"✓ Unit Count mapped: {unit_count_item['qty']} {unit_count_item['uom']}")

        # Check UOM canonicalization
        corridors_section = next((s for s in result['sections'] if s['name'] == 'Corridors'), None)
        if corridors_section:
            base_item = next((i for i in corridors_section['items'] if 'Base' in i['key']), None)
            if base_item:
                assert base_item['uom'] == 'FT', "LF should be canonicalized to FT"
                assert base_item['uom_raw'] == 'LF', "Original UOM should be preserved"
                print(f"✓ UOM canonicalization: {base_item['uom_raw']} → {base_item['uom']}")

        # Print sample raw_rows for inspection
        print("\n5. Sample raw_rows (first 3):")
        for row in result['raw_rows'][:3]:
            print(f"  Row {row['row']}: A={row.get('A')}, B={row.get('B')}, C={row.get('C')}")

        print("\n" + "="*60)
        print("✅ All API tests passed for Baycrest format!")
        print("="*60)

    finally:
        # Clean up temp file
        if os.path.exists(test_file):
            os.unlink(test_file)

        # Kill test server
        print("\nStopping test server...")
        os.system("pkill -f 'uvicorn.*8006' 2>/dev/null || true")


if __name__ == "__main__":
    test_baycrest_api()