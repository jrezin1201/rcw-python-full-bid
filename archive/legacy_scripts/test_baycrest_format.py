#!/usr/bin/env python3
"""
Test Baycrest format extraction.
Creates a synthetic Excel file matching the Baycrest/Apartment Takeoffs format
and tests the extraction process.
"""

import os
import sys
import json
import tempfile
from pathlib import Path
import openpyxl
from openpyxl import Workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from app.services.baycrest_normalizer import BaycrestNormalizer
from app.services.takeoff_mapper import TakeoffMapper


def create_test_excel(file_path: str):
    """Create a synthetic Excel file in Baycrest format."""
    wb = Workbook()

    # Create "1 Bldg" sheet
    ws = wb.active
    ws.title = "1 Bldg"

    # Header row
    ws['A1'] = 'Section'
    ws['B1'] = 'Item'
    ws['C1'] = 'Count'
    ws['D1'] = 'Ave SF'
    ws['E1'] = 'Notes'

    # General section
    ws['A2'] = 'General'
    ws['B2'] = 'Unit Count'
    ws['C2'] = 120
    ws['D2'] = 850
    ws['E2'] = 'Total unit count'

    ws['B3'] = 'Gross Building SF'
    ws['C3'] = 102000
    ws['D3'] = None
    ws['E3'] = 'Total building area'

    # Corridors section
    ws['A4'] = 'Corridors'
    ws['B4'] = 'Flooring SF'
    ws['C4'] = 8500
    ws['D4'] = None
    ws['E4'] = 'Corridor Notes'

    ws['B5'] = 'Wall SF'
    ws['C5'] = 12000

    ws['B6'] = 'Base LF'
    ws['C6'] = 1500

    # Exterior section
    ws['A7'] = 'Exterior'
    ws['B7'] = 'Stucco Wall SF'
    ws['C7'] = 25000
    ws['E7'] = 'Exterior Notes'

    ws['B8'] = 'Parapet LF'
    ws['C8'] = 450

    ws['B9'] = 'Windows'
    ws['C9'] = 240

    # Units section
    ws['A10'] = 'Units'
    ws['B10'] = 'Flooring SF'
    ws['C10'] = 85000
    ws['E10'] = 'Unit Notes'

    ws['B11'] = 'Unit Doors'
    ws['C11'] = 480

    ws['B12'] = 'Kitchen Cabinet LF'
    ws['C12'] = 3600

    # Common Areas section
    ws['A13'] = 'Common Areas'
    ws['B13'] = 'Flooring SF'
    ws['C13'] = 3500

    # Sheet / Schedule row (should be ignored)
    ws['A14'] = 'Sheet / Schedule'
    ws['B14'] = 'Reference'
    ws['C14'] = 'Page 1'

    # Empty row (should be skipped)
    ws['A15'] = None
    ws['B15'] = None
    ws['C15'] = None

    # Save the workbook
    wb.save(file_path)
    print(f"Created test Excel file: {file_path}")


def test_baycrest_extraction():
    """Test the Baycrest format extraction."""

    # Create temp file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        test_file = tmp.name

    try:
        # Create test Excel file
        create_test_excel(test_file)

        # Test normalization
        print("\n" + "="*60)
        print("Testing Baycrest Normalizer")
        print("="*60)

        normalizer = BaycrestNormalizer()
        result = normalizer.normalize_file(test_file)

        # Check raw_rows
        print(f"\nRaw rows captured: {len(result['raw_rows'])}")
        assert len(result['raw_rows']) > 0, "Should have raw_rows"

        # Print first few raw_rows for inspection
        print("\nFirst 3 raw_rows:")
        for row in result['raw_rows'][:3]:
            print(f"  Row {row['row']}: A={row['A']}, B={row['B']}, C={row['C']}, D={row['D']}, E={row['E']}")

        # Check raw_data
        print(f"\nRaw data extracted: {len(result['raw_data'])} rows")
        assert len(result['raw_data']) > 0, "Should have raw_data"

        # Print raw_data items
        print("\nExtracted raw_data:")
        for item in result['raw_data'][:5]:
            measures_str = ", ".join([f"{m['value']} {m['uom'] or 'N/A'}" for m in item['measures']])
            print(f"  [{item['section']}] {item['classification']}: {measures_str}")

        # Check extraction stats
        print(f"\nExtraction stats:")
        for key, value in result['stats'].items():
            print(f"  {key}: {value}")

        assert result['stats']['rows_total'] == len(result['raw_rows']), "rows_total should match raw_rows count"
        assert result['stats']['rows_extracted'] == len(result['raw_data']), "rows_extracted should match raw_data count"

        # Test specific extractions
        print("\n" + "="*60)
        print("Testing Specific Extractions")
        print("="*60)

        # Find Unit Count
        unit_count = next((item for item in result['raw_data']
                          if item['classification'] == 'Unit Count'), None)
        assert unit_count is not None, "Should find Unit Count"
        assert unit_count['section'] == 'General', "Unit Count should be in General section"
        assert unit_count['measures'][0]['value'] == 120, "Unit Count should be 120"
        assert unit_count['measures'][0]['uom'] == 'EA', "Unit Count should have EA uom"
        print(f"✓ Unit Count: {unit_count['measures'][0]['value']} {unit_count['measures'][0]['uom']}")

        # Find Stucco Wall SF
        stucco = next((item for item in result['raw_data']
                      if item['classification'] == 'Stucco Wall SF'), None)
        assert stucco is not None, "Should find Stucco Wall SF"
        assert stucco['section'] == 'Exterior', "Stucco should be in Exterior section"
        assert stucco['measures'][0]['value'] == 25000, "Stucco should be 25000"
        assert stucco['measures'][0]['uom'] == 'SF', "Stucco should have SF uom"
        print(f"✓ Stucco Wall: {stucco['measures'][0]['value']} {stucco['measures'][0]['uom']}")

        # Find Base LF
        base = next((item for item in result['raw_data']
                    if item['classification'] == 'Base LF'), None)
        assert base is not None, "Should find Base LF"
        assert base['measures'][0]['uom'] == 'LF', "Base should have LF uom"
        print(f"✓ Base LF: {base['measures'][0]['value']} {base['measures'][0]['uom']}")

        # Test mapping
        print("\n" + "="*60)
        print("Testing Mapping to Sections")
        print("="*60)

        mapper = TakeoffMapper(template='baycrest_v1')
        mapping_result = mapper.map_rows_to_sections(result['raw_data'])

        print(f"\nMapped sections: {len(mapping_result['sections'])}")
        for section in mapping_result['sections']:
            print(f"  {section['name']}: {len(section['items'])} items")
            for item in section['items'][:2]:  # Show first 2 items
                print(f"    - {item['key']}: {item['qty']} {item['uom']}")

        print(f"\nUnmapped items: {len(mapping_result['unmapped'])}")
        for item in mapping_result['unmapped'][:3]:
            measures_str = ", ".join([f"{m['value']} {m['uom']}" for m in item['measures']])
            print(f"  {item['classification']}: {measures_str}")

        # Check QA stats
        print(f"\nQA Stats:")
        assert 'stats' in mapping_result['qa'], "QA should have stats"
        for key, value in mapping_result['qa']['stats'].items():
            print(f"  {key}: {value}")

        # Check UOM canonicalization (canonical set: EA, SF, LF, LVL - never FT)
        print("\n" + "="*60)
        print("Testing UOM Canonicalization")
        print("="*60)

        # Find an item with LF - should stay as LF (canonical), never become FT
        corridors_section = next((s for s in mapping_result['sections'] if s['name'] == 'Corridors'), None)
        if corridors_section:
            base_item = next((i for i in corridors_section['items'] if 'Base' in i['key']), None)
            if base_item:
                print(f"Base item UOM: {base_item['uom']} (from {base_item['uom_raw']})")
                assert base_item['uom'] == 'LF', "LF should remain LF (canonical set never uses FT)"
                assert base_item['uom_raw'] == 'LF', "Original UOM should be LF"
                print("✓ UOM canonicalization working correctly (LF stays LF, not FT)")

        print("\n" + "="*60)
        print("✅ All tests passed!")
        print("="*60)

    finally:
        # Clean up temp file
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\nCleaned up test file: {test_file}")


if __name__ == "__main__":
    test_baycrest_extraction()