#!/usr/bin/env python3
"""Quick test to verify Baycrest format response structure."""
import requests
import time
import json
import tempfile
from openpyxl import Workbook

# Create test file
def create_test_file(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "1 Bldg"
    ws['A1'] = 'Section'
    ws['B1'] = 'Item'
    ws['C1'] = 'Qty'
    ws['A2'] = 'General'
    ws['B2'] = 'Unit Count'
    ws['C2'] = 100
    wb.save(path)

# Test API
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    create_test_file(tmp.name)

    # Upload
    with open(tmp.name, 'rb') as f:
        files = {'file': ('test.xlsx', f)}
        data = {'template': 'baycrest_v1'}
        headers = {'X-API-Key': 'test-api-key'}

        response = requests.post("http://127.0.0.1:8007/api/v1/jobs/",
                                 files=files, data=data, headers=headers)

    if response.status_code == 200:
        job_id = response.json()['job_id']
        print(f"Job created: {job_id}")

        # Poll
        for _ in range(10):
            time.sleep(1)
            status_response = requests.get(
                f"http://127.0.0.1:8007/api/v1/jobs/{job_id}",
                headers={'X-API-Key': 'test-api-key'}
            )

            if status_response.status_code == 200:
                result = status_response.json()
                print(f"Status: {result['status']}")

                if result['status'] == 'SUCCEEDED':
                    print("\nResponse keys:")
                    print(f"  Top level: {list(result.keys())}")
                    if result.get('result'):
                        print(f"  Result keys: {list(result['result'].keys())}")
                        if 'raw_rows' in result['result']:
                            print(f"  ✓ raw_rows present: {len(result['result']['raw_rows'])} rows")
                        else:
                            print("  ✗ raw_rows NOT present")
                        if 'raw_data' in result['result']:
                            print(f"  ✓ raw_data present: {len(result['result']['raw_data'])} rows")
                        else:
                            print("  ✗ raw_data NOT present")
                    if result.get('qa'):
                        print(f"  QA keys: {list(result['qa'].keys())}")
                        if 'stats' in result['qa']:
                            print(f"  ✓ QA stats present")
                    break
                elif result['status'] == 'FAILED':
                    print(f"Failed: {result.get('error')}")
                    break
    else:
        print(f"Upload failed: {response.status_code}")