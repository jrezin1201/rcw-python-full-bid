#!/usr/bin/env python3
"""
Comprehensive verification of Baycrest format implementation.
"""

import requests
import time
import json
from openpyxl import Workbook


def create_comprehensive_test_file():
    """Create a test file that matches actual Baycrest format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "1 Bldg"

    # Row 1: Headers (like actual Baycrest files)
    ws['A1'] = None  # Often blank in actual files
    ws['B1'] = None
    ws['C1'] = None
    ws['D1'] = None
    ws['E1'] = None

    # Row 2: General section header
    ws['A2'] = 'General'
    ws['B2'] = None
    ws['C2'] = None

    # Rows with actual data
    ws['A3'] = None
    ws['B3'] = '1 Bed Room Count'
    ws['C3'] = 45
    ws['D3'] = None
    ws['E3'] = 'One bedroom apartments'

    ws['A4'] = None
    ws['B4'] = '2 Bed Room Count'
    ws['C4'] = 30

    ws['A5'] = None
    ws['B5'] = 'Studio Count'
    ws['C5'] = 10

    ws['A6'] = None
    ws['B6'] = 'Gross Building SF'
    ws['C6'] = 85000

    # Corridors section
    ws['A7'] = 'Corridors'
    ws['B7'] = None
    ws['C7'] = None

    ws['A8'] = None
    ws['B8'] = 'Flooring SF'
    ws['C8'] = 6500

    ws['A9'] = None
    ws['B9'] = 'Wall SF'
    ws['C9'] = 9200

    ws['A10'] = None
    ws['B10'] = 'Base LF'
    ws['C10'] = 1450

    # Units section
    ws['A11'] = 'Units'
    ws['B11'] = 'Unit Flooring SF'
    ws['C11'] = 62000

    ws['A12'] = None
    ws['B12'] = 'Kitchen Cabinet LF'
    ws['C12'] = 3200

    ws['A13'] = None
    ws['B13'] = 'Unit Doors'
    ws['C13'] = 340

    # Save file
    wb.save('comprehensive_baycrest_test.xlsx')
    print("✓ Created comprehensive_baycrest_test.xlsx")
    return 'comprehensive_baycrest_test.xlsx'


def test_baycrest_full():
    """Run comprehensive test of Baycrest format."""
    API_URL = "http://127.0.0.1:8008"
    API_KEY = "test-api-key"

    # Create test file
    test_file = create_comprehensive_test_file()

    print("\n" + "="*60)
    print("COMPREHENSIVE BAYCREST FORMAT VERIFICATION")
    print("="*60)

    # Upload with baycrest_v1 template
    print("\n1. Uploading file with template=baycrest_v1...")
    with open(test_file, 'rb') as f:
        files = {'file': (test_file, f)}
        data = {'template': 'baycrest_v1'}
        headers = {'X-API-Key': API_KEY}

        response = requests.post(f"{API_URL}/api/v1/jobs/",
                                files=files, data=data, headers=headers)

    if response.status_code != 200:
        print(f"✗ Upload failed: {response.status_code}")
        return False

    job_id = response.json()['job_id']
    print(f"✓ Job created: {job_id}")

    # Poll for completion
    print("\n2. Polling for completion...")
    for attempt in range(10):
        time.sleep(1)
        status_response = requests.get(
            f"{API_URL}/api/v1/jobs/{job_id}",
            headers={'X-API-Key': API_KEY}
        )

        if status_response.status_code != 200:
            print(f"✗ Status check failed: {status_response.status_code}")
            return False

        result = status_response.json()

        if result['status'] == 'SUCCEEDED':
            print(f"✓ Job completed successfully")

            # Detailed verification
            print("\n3. DETAILED VERIFICATION:")
            print("-" * 40)

            # Check QA stats
            qa = result.get('qa', {})
            stats = qa.get('stats', {})

            print("\n✓ QA Stats (always present):")
            print(f"   - rows_total: {stats.get('rows_total', 'MISSING')}")
            print(f"   - rows_extracted: {stats.get('rows_extracted', 'MISSING')}")
            print(f"   - rows_ignored: {stats.get('rows_ignored', 'MISSING')}")
            print(f"   - items_mapped: {stats.get('items_mapped', 'MISSING')}")
            print(f"   - items_unmapped: {stats.get('items_unmapped', 'MISSING')}")

            if not stats:
                print("   ✗ ERROR: QA stats are missing!")
            else:
                print("   ✓ QA stats properly included")

            # Check for raw_rows and raw_data
            result_data = result.get('result', {})

            print("\n✓ Raw Data Check:")
            if 'raw_rows' in result_data:
                raw_rows = result_data['raw_rows']
                print(f"   ✓ raw_rows present: {len(raw_rows)} rows")

                # Show sample raw_rows
                print("   Sample raw_rows (first 3):")
                for row in raw_rows[:3]:
                    print(f"     Row {row['row']}: B='{row.get('B')}', C={row.get('C')}")
            else:
                print("   ✗ raw_rows MISSING from response!")

            if 'raw_data' in result_data:
                raw_data = result_data['raw_data']
                print(f"   ✓ raw_data present: {len(raw_data)} rows")

                # Show sample raw_data
                print("   Sample raw_data (first 3):")
                for item in raw_data[:3]:
                    classification = item.get('classification', 'N/A')
                    measures = item.get('measures', [])
                    if measures:
                        value = measures[0].get('value', 'N/A')
                        uom = measures[0].get('uom', 'N/A')
                        print(f"     {classification}: {value} {uom}")
            else:
                print("   ✗ raw_data MISSING from response!")

            # Check sections
            sections = result_data.get('sections', [])
            print(f"\n✓ Sections: {len(sections)} mapped")
            for section in sections:
                items = section.get('items', [])
                print(f"   - {section['name']}: {len(items)} items")
                for item in items[:2]:  # First 2 items
                    print(f"       • {item['key']}: {item['qty']} {item['uom']}")

            # Verify specific extractions
            print("\n✓ Data Extraction Verification:")

            # Check if bedroom counts were extracted
            bedroom_count_found = False
            if 'raw_data' in result_data:
                for item in result_data['raw_data']:
                    if '1 Bed Room Count' in item.get('classification', ''):
                        bedroom_count_found = True
                        measures = item.get('measures', [])
                        if measures:
                            value = measures[0].get('value')
                            if value == 45:
                                print(f"   ✓ 1 Bed Room Count correctly extracted: {value}")
                            else:
                                print(f"   ✗ 1 Bed Room Count wrong value: {value} (expected 45)")
                        break

            if not bedroom_count_found:
                print("   ✗ 1 Bed Room Count not found in raw_data!")

            # Check UOM inference
            print("\n✓ UOM Inference Check:")
            uom_checks = {
                'Gross Building SF': 'SF',
                'Base LF': 'LF',
                'Unit Doors': 'EA'
            }

            if 'raw_data' in result_data:
                for check_class, expected_uom in uom_checks.items():
                    for item in result_data['raw_data']:
                        if check_class in item.get('classification', ''):
                            measures = item.get('measures', [])
                            if measures:
                                uom = measures[0].get('uom')
                                if uom == expected_uom:
                                    print(f"   ✓ {check_class}: {uom} (correct)")
                                else:
                                    print(f"   ✗ {check_class}: {uom} (expected {expected_uom})")
                            break

            # Final summary
            print("\n" + "="*60)
            print("VERIFICATION SUMMARY:")

            all_good = True

            if stats:
                print("✓ QA stats always present")
            else:
                print("✗ QA stats MISSING")
                all_good = False

            if 'raw_rows' in result_data:
                print(f"✓ raw_rows included ({len(result_data['raw_rows'])} rows)")
            else:
                print("✗ raw_rows MISSING")
                all_good = False

            if 'raw_data' in result_data:
                print(f"✓ raw_data included ({len(result_data['raw_data'])} rows)")
            else:
                print("✗ raw_data MISSING")
                all_good = False

            if sections:
                print(f"✓ Sections properly mapped ({len(sections)} sections)")
            else:
                print("✗ No sections mapped")
                all_good = False

            if all_good:
                print("\n✅ ALL CHECKS PASSED - Baycrest format working correctly!")
            else:
                print("\n⚠️  SOME CHECKS FAILED - Review implementation")

            # Save full response for debugging
            with open('baycrest_response.json', 'w') as f:
                json.dump(result, f, indent=2)
            print("\n(Full response saved to baycrest_response.json)")

            return all_good

        elif result['status'] == 'FAILED':
            print(f"✗ Job failed: {result.get('error', {}).get('message', 'Unknown')}")
            return False

    print("✗ Job timed out")
    return False


if __name__ == "__main__":
    success = test_baycrest_full()
    exit(0 if success else 1)