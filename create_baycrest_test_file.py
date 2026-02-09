#!/usr/bin/env python3
"""
Create a Baycrest format Excel file for testing.
This matches the actual "Apartment Takeoffs - Baycrest" format.
"""

from openpyxl import Workbook

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "1 Bldg"

# Row 1: Headers
ws['A1'] = 'Section'
ws['B1'] = 'Classification'
ws['C1'] = 'Qty'
ws['D1'] = 'Secondary'
ws['E1'] = 'Notes'

# Row 2: General section
ws['A2'] = 'General'
ws['B2'] = None
ws['C2'] = None

# Row 3: Actual data
ws['A3'] = None
ws['B3'] = '1 Bed Room Count'
ws['C3'] = 35
ws['D3'] = None
ws['E3'] = 'One bedroom units'

# Row 4: More data
ws['B4'] = '2 Bed Room Count'
ws['C4'] = 25
ws['D4'] = None

# Row 5: Studio
ws['B5'] = 'Studio Count'
ws['C5'] = 10
ws['D5'] = None

# Row 6: Building SF
ws['B6'] = 'Gross Building SF'
ws['C6'] = 75000
ws['D6'] = None

# Row 7: Corridors section
ws['A7'] = 'Corridors'
ws['B7'] = 'Flooring SF'
ws['C7'] = 5500
ws['D7'] = None

# Row 8: More corridor data
ws['B8'] = 'Wall SF'
ws['C8'] = 8200
ws['D8'] = None

# Row 9: Base
ws['B9'] = 'Base LF'
ws['C9'] = 1200
ws['D9'] = None

# Row 10: Units section
ws['A10'] = 'Units'
ws['B10'] = 'Unit Flooring SF'
ws['C10'] = 52000
ws['D10'] = None

# Row 11: Kitchen
ws['B11'] = 'Kitchen Cabinet LF'
ws['C11'] = 2800
ws['D11'] = None

# Row 12: Doors
ws['B12'] = 'Unit Doors'
ws['C12'] = 280
ws['D12'] = None

# Row 13: Empty row (should be ignored)
ws['A13'] = None
ws['B13'] = None
ws['C13'] = None

# Row 14: Exterior section
ws['A14'] = 'Exterior'
ws['B14'] = 'Stucco Wall SF'
ws['C14'] = 18000
ws['D14'] = None

# Row 15: Windows
ws['B15'] = 'Windows'
ws['C15'] = 150
ws['D15'] = None

# Save the file
wb.save('baycrest_test.xlsx')
print("Created baycrest_test.xlsx in Baycrest format")
print("\nKey characteristics:")
print("  - Sheet name: '1 Bldg'")
print("  - Column A: Section headers (General, Corridors, Units, Exterior)")
print("  - Column B: Classifications (1 Bed Room Count, Flooring SF, etc.)")
print("  - Column C: Numeric quantities")
print("  - Column D: Optional secondary values")
print("  - Column E: Notes")
print("\nTotal data rows: 12 (should extract 12 items)")