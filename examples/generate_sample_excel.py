#!/usr/bin/env python3
"""
Generate sample Excel file for testing the extraction service.
Creates a realistic bid/invoice spreadsheet with various data types and edge cases.
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_sample_excel():
    """Create a sample Excel file with various test cases."""

    # Create workbook and get active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Items"

    # Set column widths
    column_widths = {
        'A': 15,  # Item Code
        'B': 40,  # Description
        'C': 12,  # Quantity
        'D': 10,  # Unit
        'E': 15,  # Unit Price
        'F': 15,  # Total Price
        'G': 25,  # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add company header (merged cells)
    ws.merge_cells('A1:G1')
    ws['A1'] = "R.C. WENDT & ASSOCIATES"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:G2')
    ws['A2'] = "Construction Materials Bid - Project #2024-001"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Add date
    ws['A3'] = "Date:"
    ws['B3'] = datetime.now().strftime("%m/%d/%Y")

    # Skip a row
    header_row = 5

    # Add headers
    headers = [
        "Item Code",
        "Description",
        "Quantity",
        "Unit",
        "Unit Price",
        "Total Price",
        "Notes"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    # Sample data with various edge cases
    items = [
        # Normal items
        ("MAT-001", "Concrete Mix - Type III Portland Cement", 150, "CY", 125.50, "Standard mix"),
        ("MAT-002", "Reinforcing Steel - #4 Rebar", 2500, "LF", 2.75, "Grade 60"),
        ("MAT-003", "Structural Steel - W12x26 Beam", 35, "EA", 450.00, "A992 Grade 50"),

        # Item with merged description (simulating wrapped text)
        ("MAT-004", "Aggregate Base Course\nClass 5, Compacted", 500, "TON", 18.50, "3/4 inch minus"),

        # Items with different number formats
        ("LAB-001", "Skilled Labor - Ironworker", 80, "HR", 85.00, "Prevailing wage"),
        ("LAB-002", "Equipment Operator", 40.5, "HR", 75.50, "Overtime included"),

        # Item with percentage (potential anomaly)
        ("MISC-001", "Contingency", 1, "LS", "5%", "Percentage of total"),

        # Large numbers (scientific notation risk)
        ("MAT-005", "Sand Fill", 10000, "CY", 12.00, "Delivered"),

        # Empty quantity (should be flagged)
        ("MAT-006", "Geotextile Fabric", "", "SY", 3.25, "Type not specified"),

        # Text in quantity field (anomaly)
        ("MAT-007", "Custom Fabrication", "TBD", "EA", 500.00, "Requires shop drawings"),

        # Very small numbers
        ("CHM-001", "Epoxy Adhesive", 0.5, "GAL", 125.00, "Two-part system"),

        # Missing price
        ("MAT-008", "Owner Supplied Material", 100, "EA", "", "No charge - owner supplied"),
    ]

    # Add data rows
    data_start_row = header_row + 1
    for row_idx, item in enumerate(items, start=data_start_row):
        code, desc, qty, unit, price, notes = item

        # Item code
        ws.cell(row=row_idx, column=1, value=code).border = border_style

        # Description
        ws.cell(row=row_idx, column=2, value=desc).border = border_style
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

        # Quantity
        qty_cell = ws.cell(row=row_idx, column=3)
        qty_cell.value = qty
        qty_cell.border = border_style
        if isinstance(qty, (int, float)):
            qty_cell.number_format = '#,##0.00'

        # Unit
        ws.cell(row=row_idx, column=4, value=unit).border = border_style

        # Unit Price
        price_cell = ws.cell(row=row_idx, column=5)
        price_cell.value = price
        price_cell.border = border_style
        if isinstance(price, (int, float)):
            price_cell.number_format = '$#,##0.00'

        # Total Price (formula)
        total_cell = ws.cell(row=row_idx, column=6)
        if isinstance(qty, (int, float)) and isinstance(price, (int, float)):
            total_cell.value = f"=C{row_idx}*E{row_idx}"
            total_cell.number_format = '$#,##0.00'
        else:
            total_cell.value = "N/A"
        total_cell.border = border_style

        # Notes
        ws.cell(row=row_idx, column=7, value=notes).border = border_style

    # Add subtotal row (potential totals row for detection)
    total_row = data_start_row + len(items) + 1
    ws.merge_cells(f'A{total_row}:B{total_row}')
    ws[f'A{total_row}'] = "SUBTOTAL"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right")

    # Add total formula
    ws[f'F{total_row}'] = f"=SUM(F{data_start_row}:F{total_row-1})"
    ws[f'F{total_row}'].number_format = '$#,##0.00'
    ws[f'F{total_row}'].font = Font(bold=True)

    # Add some empty rows (should be filtered out)
    for i in range(3):
        total_row += 1
        # Empty row - no data

    # Add grand total
    total_row += 1
    ws.merge_cells(f'A{total_row}:B{total_row}')
    ws[f'A{total_row}'] = "GRAND TOTAL"
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right")

    # Create second sheet with different format
    ws2 = wb.create_sheet("Equipment List")

    # Simple header without merged cells
    ws2['A1'] = "Equipment ID"
    ws2['B1'] = "Equipment Name"
    ws2['C1'] = "Daily Rate"
    ws2['D1'] = "Status"

    # Style the headers
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Add equipment data
    equipment = [
        ("EQ-001", "Excavator CAT 320", 1200.00, "Available"),
        ("EQ-002", "Bulldozer D8", 1500.00, "In Use"),
        ("EQ-003", "Crane 50 Ton", 2500.00, "Available"),
        ("EQ-004", "Dump Truck", 800.00, "Maintenance"),
    ]

    for row_idx, eq in enumerate(equipment, start=2):
        for col_idx, value in enumerate(eq, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3:  # Daily rate column
                cell.number_format = '$#,##0.00'

    # Save the workbook
    output_path = Path(__file__).parent / "sample_bid.xlsx"
    wb.save(output_path)
    print(f"Sample Excel file created: {output_path}")

    return output_path


if __name__ == "__main__":
    create_sample_excel()