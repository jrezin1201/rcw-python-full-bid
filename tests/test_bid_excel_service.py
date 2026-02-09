from pathlib import Path

from openpyxl import load_workbook

from app.services.bid_excel_service import (
    INTERNAL_MARKER,
    export_internal_bid_workbook,
    export_proposal_workbook,
    import_internal_bid_workbook,
    is_internal_bid_workbook,
)
from app.ui.excel_mapper import map_excel_with_catalog


def test_internal_export_import_roundtrip(tmp_path):
    state, _, _ = map_excel_with_catalog(
        str(Path(__file__).parent / "test_data" / "client_input_data.xlsx"),
        template="baycrest_v1",
    )
    content = export_internal_bid_workbook(state)
    out_path = tmp_path / "internal.xlsx"
    out_path.write_bytes(content)

    assert is_internal_bid_workbook(str(out_path))

    imported = import_internal_bid_workbook(str(out_path))
    assert imported.total_items == state.total_items
    assert imported.project_name == state.project_name


def test_internal_export_contains_marker(tmp_path):
    state, _, _ = map_excel_with_catalog(
        str(Path(__file__).parent / "test_data" / "client_input_data.xlsx"),
        template="baycrest_v1",
    )
    content = export_internal_bid_workbook(state)
    out_path = tmp_path / "internal.xlsx"
    out_path.write_bytes(content)

    wb = load_workbook(out_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    assert ws["A1"].value == INTERNAL_MARKER
    wb.close()


def test_proposal_export_writes_pricing_rows(tmp_path):
    state, _, _ = map_excel_with_catalog(
        str(Path(__file__).parent / "test_data" / "client_input_data.xlsx"),
        template="baycrest_v1",
    )
    # Ensure alternate block gets populated from state.
    state.items[0].is_alternate = True
    state.items[0].name = "Alt Test Item"
    state.items[0].unit_price_base = 10
    state.items[0].qty = 2
    content = export_proposal_workbook(state)
    out_path = tmp_path / "proposal.xlsx"
    out_path.write_bytes(content)

    wb = load_workbook(out_path, data_only=False)
    ws = wb["Sheet1"]
    # Pricing table in proposal template
    assert ws["E98"].value is not None
    assert ws["E105"].value is not None
    assert isinstance(ws["E105"].value, (int, float))
    assert ws["A121"].value == "Alt Test Item"
    assert ws["D121"].value == 20
    wb.close()
