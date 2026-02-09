"""
Generate synthetic Excel files for testing takeoff extraction.
"""
from pathlib import Path

import openpyxl
from openpyxl import Workbook


def create_standard_takeoff():
    """Create a standard takeoff Excel file with expected format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Takeoff Data"

    # Headers
    headers = [
        "Classification",
        "Quantity",
        "Quantity1 UOM",
        "Quantity2",
        "Quantity2 UOM",
        "Quantity3",
        "Quantity3 UOM"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Test data with various cases
    test_data = [
        # Standard items with exact matches
        ["Unit Count", 14, "EA", None, None, None, None],
        ["Total SF", 25000, "SF", None, None, None, None],
        ["Cor. Lid", 3500, "SF", None, None, None, None],

        # Item with multiple measures
        ["Kitchen Cabinet", 150, "LF", 200, "SF", 50, "EA"],

        # Fuzzy match cases
        ["Units Count", 15, "EA", None, None, None, None],  # Slight variation
        ["Corridor Ceiling", 3600, "SF", None, None, None, None],  # Different name

        # Items with FT that should convert to LF
        ["Cor. Base", 1200, "FT", None, None, None, None],

        # Items with commas in numbers
        ["Unit Wall", "12,500", "SF", None, None, None, None],

        # Unknown classification
        ["Random Item XYZ", 100, "EA", None, None, None, None],

        # Multiple measures with same UOM (should pick largest)
        ["Window", 10, "EA", 15, "EA", 5, "EA"],

        # Empty row to be skipped
        [None, None, None, None, None, None, None],

        # More realistic items
        ["Interior Door", 85, "EA", None, None, None, None],
        ["Paint", 45000, "SF", None, None, None, None],
        ["Carpet", 8500, "SF", None, None, None, None],
        ["Toilet", 28, "EA", None, None, None, None],
    ]

    # Write data
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Save file
    output_path = Path(__file__).parent / "standard_takeoff.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_edge_case_takeoff():
    """Create a takeoff with edge cases and problematic data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Problem Data"

    # Headers with slightly different naming
    headers = [
        "Class",  # Different header name
        "Qty",
        "UOM1",
        "Qty2",
        "UOM2",
        "Quantity3",
        "Quantity3 UOM"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Edge case data
    test_data = [
        # Missing classification
        ["", 100, "SF", None, None, None, None],

        # Text in quantity field
        ["Unit Count", "TBD", "EA", None, None, None, None],

        # Missing UOM
        ["Total SF", 5000, None, None, None, None, None],

        # Very large number
        ["Paint", 1234567.89, "SF", None, None, None, None],

        # Special characters
        ["Unit Wall (Type A)", 3000, "SF", None, None, None, None],

        # All caps
        ["CORRIDOR CEILING", 2500, "SQFT", None, None, None, None],

        # Mixed case UOM
        ["Kitchen Cabinet", 100, "lf", None, None, None, None],

        # Multiple spaces
        ["  Unit  Floor  ", 8000, " SF ", None, None, None, None],
    ]

    # Write data
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Save file
    output_path = Path(__file__).parent / "edge_case_takeoff.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_minimal_takeoff():
    """Create a minimal valid takeoff file."""
    wb = Workbook()
    ws = wb.active

    # Simple headers
    ws['A1'] = "Classification"
    ws['B1'] = "Quantity"
    ws['C1'] = "Quantity1 UOM"

    # Just a few rows
    ws['A2'] = "Unit Count"
    ws['B2'] = 10
    ws['C2'] = "EA"

    ws['A3'] = "Total SF"
    ws['B3'] = 5000
    ws['C3'] = "SF"

    # Save file
    output_path = Path(__file__).parent / "minimal_takeoff.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_standard_takeoff()
    create_edge_case_takeoff()
    create_minimal_takeoff()
    print("Test files generated successfully!")